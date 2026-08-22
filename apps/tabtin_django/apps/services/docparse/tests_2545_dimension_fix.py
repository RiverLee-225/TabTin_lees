"""
#2545 回归测试 —— read_only xlsx 解析不再被过小 <dimension> 截成 1×1。

根因：openpyxl read_only 模式下 iter_rows 范围取自 sheet XML 的 <dimension>
元素；部分生产者写过小 dimension（如 ref="A1"），实际数据被静默截断。
修复：_read_sheet_rows 入口调 ws.reset_dimensions() 清掉声明 dimension，
iter_rows 按实际单元格全扫。

fixture 内联生成（本分支从 release 签出，没有 harness 分支的
apps/tabtin-electron/fixtures/poc-xlsx/；生成手法与 harness 分支
scripts/xlsx-2545-build-fixtures.py 同款：openpyxl 写正常文件 + zip 后处理改坏
<dimension>）。harness 分支合入后两套 fixture 并存，互不冲突。

复跑：
  USE_SQLITE_FOR_TESTS=0 python apps/tabtin_django/manage.py test \\
      apps.services.docparse.tests_2545_dimension_fix
"""
from __future__ import annotations

import os
import re
import shutil
import tempfile
import zipfile

from django.test import SimpleTestCase

from apps.services.docparse.parsers.xlsx_parser import ExcelParser


def _write_3x4_xlsx(path: str) -> None:
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for r in range(1, 4):
        for c in range(1, 5):
            ws.cell(row=r, column=c, value=f"r{r}c{c}")
    wb.save(path)


def _corrupt_dimension(xlsx_path: str, new_ref: str) -> None:
    """把 sheet XML 的 <dimension ref="..."/> 改成过小的 new_ref（如 A1）。"""
    tmp_path = xlsx_path + ".tmp"
    with zipfile.ZipFile(xlsx_path, "r") as zin, zipfile.ZipFile(
        tmp_path, "w", zipfile.ZIP_DEFLATED
    ) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename.startswith("xl/worksheets/sheet") and item.filename.endswith(".xml"):
                text = data.decode("utf-8")
                text = re.sub(
                    r'<dimension\s+ref="[^"]*"\s*/>',
                    f'<dimension ref="{new_ref}"/>',
                    text,
                )
                data = text.encode("utf-8")
            zout.writestr(item, data)
    shutil.move(tmp_path, xlsx_path)


class Xlsx2545DimensionFixTest(SimpleTestCase):
    """wrong dimension 不再截断；正常文件 / 空 sheet 行为不回归。"""

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.tmpdir = tempfile.mkdtemp(prefix="tabtin-2545-")

    @classmethod
    def tearDownClass(cls):
        shutil.rmtree(cls.tmpdir, ignore_errors=True)
        super().tearDownClass()

    def _dims(self, path: str) -> list[tuple[int, int]]:
        result = ExcelParser().parse(path)
        out = []
        for page in result.pages:
            md = page.chunks[0].metadata if page.chunks else {}
            out.append((md.get("rows", 0), md.get("cols", 0)))
        return out

    def test_wrong_dimension_parses_full_3x4(self):
        """核心回归：<dimension ref="A1"/> 但实际 3×4 → 应解析出 3×4（修复前 1×1）。"""
        path = os.path.join(self.tmpdir, "wrong_dimension.xlsx")
        _write_3x4_xlsx(path)
        _corrupt_dimension(path, "A1")
        self.assertEqual(self._dims(path), [(3, 4)])

    def test_normal_dimension_unchanged(self):
        """正确 dimension 的文件输出不变（3×4）。"""
        path = os.path.join(self.tmpdir, "normal.xlsx")
        _write_3x4_xlsx(path)
        self.assertEqual(self._dims(path), [(3, 4)])

    def test_empty_sheet_still_skipped(self):
        """空 sheet 在 reset_dimensions 后 iter_rows 产 0 行 → 仍被跳过，不产出 page。"""
        import openpyxl

        path = os.path.join(self.tmpdir, "empty.xlsx")
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "HasData"
        for c in range(1, 3):
            ws1.cell(row=1, column=c, value=f"h{c}")
        wb.create_sheet("EmptySheet")
        wb.save(path)
        result = ExcelParser().parse(path)
        self.assertEqual(len(result.pages), 1)
        self.assertEqual(result.pages[0].chunks[0].metadata["sheet_name"], "HasData")

    def test_undersized_multirow_dimension(self):
        """dimension 声明 A1:B2（小于实际 3×4）→ 也应读全 3×4。"""
        path = os.path.join(self.tmpdir, "undersized.xlsx")
        _write_3x4_xlsx(path)
        _corrupt_dimension(path, "A1:B2")
        self.assertEqual(self._dims(path), [(3, 4)])

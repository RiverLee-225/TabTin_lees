"""
Excel 解析器

将 .xlsx 文件解析为结构化 chunks：
- 每个 Sheet 视为一个"逻辑页"
- 每个 Sheet 产出一个 table 类型的 chunk（Markdown 表格）
- 仅支持 .xlsx（openpyxl），不支持旧版 .xls
"""

from __future__ import annotations

import logging
import math

from .base import BaseDocumentParser, ChunkResult, PageResult, ParseResult
from .registry import register_parser

logger = logging.getLogger(__name__)

_MAX_ROWS_PER_SHEET = 2000
_MAX_COLS_PER_SHEET = 100


@register_parser
class ExcelParser(BaseDocumentParser):

    def supported_mimes(self) -> list[str]:
        return [
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        ]

    def parse(self, file_path: str, **kwargs) -> ParseResult:
        import openpyxl

        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        pages: list[PageResult] = []
        sheet_names: list[str] = list(wb.sheetnames)

        try:
            for sheet_idx, sheet_name in enumerate(sheet_names):
                ws = wb[sheet_name]
                page_number = sheet_idx + 1

                rows = _read_sheet_rows(ws)
                if not rows:
                    continue

                md_lines = _rows_to_markdown(rows)
                md_content = "\n".join(md_lines)

                chunks = [
                    ChunkResult(
                        chunk_type="table",
                        content=md_content,
                        sequence=1,
                        metadata={
                            "source": "structural",
                            "sheet_name": sheet_name,
                            "rows": len(rows),
                            "cols": len(rows[0]) if rows else 0,
                        },
                    ),
                ]

                text_content = f"[Sheet: {sheet_name}]\n{md_content}"
                pages.append(PageResult(
                    page_number=page_number,
                    width=0,
                    height=0,
                    chunks=chunks,
                    text_content=text_content,
                ))
        finally:
            wb.close()

        title = sheet_names[0] if sheet_names else ""

        return ParseResult(
            pages=pages,
            title=title,
            parse_method="structural",
        )


def _read_sheet_rows(ws) -> list[list[str]]:
    """读取 Sheet 的所有行，截断超大表格。"""
    rows: list[list[str]] = []

    #  修复：read_only 模式下 iter_rows 的范围取自 sheet XML 的 <dimension>
    # 元素（max_row/max_column）。部分生产者写过小的 dimension（如 ref="A1"），
    # 会把实际 3×N 数据静默截成 1×1。reset_dimensions() 清掉声明的 dimension
    # （max_row/max_column → None），让 iter_rows 按实际单元格流式全扫。
    # 空 sheet 调用后 iter_rows 产 0 行，行为不变；正确 dimension 的文件输出也
    # 不变（实际单元格范围与声明一致）。normal 模式的 Worksheet 无此方法，
    # hasattr 守卫兼容单测传入的非 read_only worksheet / MagicMock 之外的替身。
    if hasattr(ws, "reset_dimensions"):
        ws.reset_dimensions()

    for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
        if row_idx >= _MAX_ROWS_PER_SHEET:
            logger.info("Sheet 行数超过 %d，截断", _MAX_ROWS_PER_SHEET)
            break

        cells = []
        for col_idx, cell in enumerate(row):
            if col_idx >= _MAX_COLS_PER_SHEET:
                break
            cells.append(_cell_to_str(cell))
        rows.append(cells)

    if not rows:
        return []

    while rows and all(c == "" for c in rows[-1]):
        rows.pop()

    return rows


def _cell_to_str(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        if math.isinf(value) or math.isnan(value):
            return ""
        if value == int(value):
            return str(int(value))
        return f"{value:.6g}"
    return str(value).replace("|", "\\|").replace("\n", " ").strip()


def _rows_to_markdown(rows: list[list[str]]) -> list[str]:
    if not rows:
        return []

    max_cols = max(len(r) for r in rows)

    def _pad(row: list[str]) -> list[str]:
        return row + [""] * (max_cols - len(row))

    header = _pad(rows[0])
    lines = ["| " + " | ".join(header) + " |"]
    lines.append("| " + " | ".join("---" for _ in header) + " |")

    for row in rows[1:]:
        padded = _pad(row)
        lines.append("| " + " | ".join(padded) + " |")

    return lines

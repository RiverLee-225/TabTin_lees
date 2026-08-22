from unittest.mock import MagicMock, call as mock_call, patch
from uuid import uuid4

from django.contrib.auth import get_user_model
from django.test import SimpleTestCase, TestCase

from apps.tabdata.api_import_export import _get_export_view_query
from apps.tabdata.models import Table, TableField, TableRecord, TableView
from apps.tabdata.schemas import ExportToExcelRequest
from apps.tabdata.services.export_service import ExportService
from apps.tabtinspace.tests.fixtures import create_test_organization_with_agent


User = get_user_model()


class FilteredViewExportPaginationTest(SimpleTestCase):
    """导出应复用当前视图的数据链路，并完整遍历匹配记录。"""

    def setUp(self):
        self.service = ExportService(user=MagicMock())
        self.view = MagicMock()

    @patch('apps.tabdata.services.view_data_service.ViewDataService._get_grid_data')
    def test_collects_all_filtered_record_ids_across_pages(self, mock_get_grid_data):
        record_ids = [str(uuid4()) for _ in range(5)]
        mock_get_grid_data.side_effect = [
            {'records': [{'id': record_id} for record_id in record_ids[:2]]},
            {'records': [{'id': record_id} for record_id in record_ids[2:4]]},
            {'records': [{'id': record_ids[4]}]},
        ]

        with patch('apps.tabdata.services.export_service._VIEW_EXPORT_PAGE_SIZE', 2, create=True):
            result = self.service._get_view_record_ids(self.view, max_rows=5)

        self.assertEqual(result, record_ids)
        self.assertEqual(
            [mock_call.kwargs['page'] for mock_call in mock_get_grid_data.call_args_list],
            [1, 2, 3],
        )

    @patch('apps.tabdata.services.view_data_service.ViewDataService._get_grid_data')
    def test_exact_boundary_page_stops_after_empty_page(self, mock_get_grid_data):
        record_ids = [str(uuid4()) for _ in range(4)]
        mock_get_grid_data.side_effect = [
            {'records': [{'id': record_id} for record_id in record_ids[:2]]},
            {'records': [{'id': record_id} for record_id in record_ids[2:]]},
            {'records': []},
        ]

        with patch('apps.tabdata.services.export_service._VIEW_EXPORT_PAGE_SIZE', 2, create=True):
            result = self.service._get_view_record_ids(self.view, max_rows=10)

        self.assertEqual(result, record_ids)
        self.assertEqual(mock_get_grid_data.call_count, 3)

    @patch('apps.tabdata.services.view_data_service.ViewDataService._get_grid_data')
    def test_empty_filtered_view_returns_no_record_ids(self, mock_get_grid_data):
        mock_get_grid_data.return_value = {'records': []}

        result = self.service._get_view_record_ids(self.view, max_rows=10)

        self.assertEqual(result, [])
        mock_get_grid_data.assert_called_once()

    @patch('apps.tabdata.services.view_data_service.ViewDataService._get_grid_data')
    def test_transient_view_query_is_forwarded_to_every_page(self, mock_get_grid_data):
        filters = [{'field_id': str(uuid4()), 'operator': 'equal', 'value': '待处理'}]
        sorts = [{'field_id': str(uuid4()), 'direction': 'desc'}]
        groups = [{'field_id': str(uuid4()), 'direction': 'asc'}]
        mock_get_grid_data.side_effect = [
            {'records': [{'id': str(uuid4())}, {'id': str(uuid4())}]},
            {'records': []},
        ]

        with patch('apps.tabdata.services.export_service._VIEW_EXPORT_PAGE_SIZE', 2, create=True):
            self.service._get_view_record_ids(
                self.view,
                max_rows=10,
                filters=filters,
                filter_logic='or',
                sorts=sorts,
                groups=groups,
            )

        self.assertEqual(mock_get_grid_data.call_count, 2)
        for grid_call in mock_get_grid_data.call_args_list:
            self.assertEqual(grid_call.kwargs['filters'], filters)
            self.assertEqual(grid_call.kwargs['filter_logic'], 'or')
            self.assertEqual(grid_call.kwargs['sorts'], sorts)
            self.assertEqual(grid_call.kwargs['groups'], groups)

    def test_excel_request_preserves_transient_view_query(self):
        filters = [{'field_id': str(uuid4()), 'operator': 'equal', 'value': '待处理'}]
        request = ExportToExcelRequest(
            table_id=uuid4(),
            view_id=uuid4(),
            filters=filters,
            filter_logic='and',
        )

        self.assertEqual(request.filters, filters)
        self.assertEqual(request.filter_logic, 'and')

    def test_empty_query_arrays_remain_explicit_overrides(self):
        request = ExportToExcelRequest(
            table_id=uuid4(),
            view_id=uuid4(),
            filters=[],
            sorts=[],
            groups=[],
        )

        self.assertEqual(
            _get_export_view_query(request),
            {'filters': [], 'sorts': [], 'groups': []},
        )

    def test_legacy_request_keeps_view_query_absent(self):
        request = ExportToExcelRequest(table_id=uuid4(), view_id=uuid4())

        self.assertEqual(_get_export_view_query(request), {})

    @patch.object(ExportService, '_get_view_record_ids', return_value=[])
    @patch.object(TableView, 'objects')
    @patch.object(TableRecord, 'objects')
    def test_empty_view_ids_never_fall_back_to_full_table(
        self,
        mock_record_objects,
        mock_view_objects,
        mock_get_view_record_ids,
    ):
        queryset = MagicMock()
        empty_queryset = queryset.none.return_value
        mock_record_objects.using.return_value.filter.return_value = queryset
        mock_view_objects.using.return_value.get.return_value = self.view

        filters = []
        result = self.service._get_records_queryset(
            uuid4(),
            view_id=uuid4(),
            filters=filters,
            filter_logic='and',
            sorts=[],
            groups=[],
        )

        self.assertIs(result, empty_queryset)
        queryset.filter.assert_not_called()
        mock_get_view_record_ids.assert_called_once_with(
            self.view,
            max_rows=None,
            rls_context=None,
            filters=filters,
            filter_logic='and',
            sorts=[],
            groups=[],
        )


class FilteredViewExportIntegrationTest(TestCase):
    """Use the real PostgreSQL view query and XLSX writer, including pagination."""

    databases = ['default', 'postgresql']

    def setUp(self):
        self.user = User.objects.create_user(
            phone='13800009991',
            nickname='filtered-export-test',
        )
        context = create_test_organization_with_agent(
            owner=self.user,
            organization_name='Filtered export organization',
            space_name='Filtered export space',
            prefix='filtered_export',
        )
        self.space = context['space']
        self.table = Table.objects.create(
            space_id=self.space.id,
            organization_id=self.space.organization_id,
            name='Filtered export table',
            owner=self.user,
        )
        self.title_field = TableField.objects.create(
            table=self.table,
            name='Title',
            field_type='text',
            is_primary=True,
            order=0,
        )
        self.status_field = TableField.objects.create(
            table=self.table,
            name='Status',
            field_type='select',
            config={
                'options': [
                    {'value': 'pending', 'label': 'Pending'},
                    {'value': 'processing', 'label': 'Processing'},
                    {'value': 'fixed', 'label': 'Fixed'},
                ]
            },
            order=1,
        )
        self.view = TableView.objects.create(
            table=self.table,
            name='Grid',
            view_type='grid',
            created_by=self.user,
            filters=[],
            config={},
        )
        statuses = ('pending',) * 6 + ('processing', 'fixed')
        for index, status in enumerate(statuses):
            TableRecord.objects.create(
                table=self.table,
                created_by=self.user,
                order=index,
                data={
                    str(self.title_field.id): f'row-{index}',
                    str(self.status_field.id): status,
                },
            )
        TableRecord.objects.create(
            table=self.table,
            created_by=self.user,
            order=len(statuses),
            data={str(self.title_field.id): 'row-8'},
        )

    def _pending_view_query(self):
        return {
            'filters': [
                {
                    'id': 'status-filter',
                    'field_id': str(self.status_field.id),
                    'operator': 'isAnyOf',
                    'value': ['pending'],
                    'enabled': True,
                }
            ],
            'filter_logic': 'and',
            'sorts': [],
            'groups': [],
        }

    @patch('apps.tabdata.services.export_service._VIEW_EXPORT_PAGE_SIZE', 2)
    def test_excel_contains_six_pending_rows_across_pages(self):
        import io

        import openpyxl

        result = ExportService(user=self.user).export_to_excel(
            table_id=self.table.id,
            view_id=self.view.id,
            field_ids=[self.title_field.id, self.status_field.id],
            include_headers=True,
            view_query=self._pending_view_query(),
        )

        worksheet = openpyxl.load_workbook(io.BytesIO(result), read_only=True).active
        rows = list(worksheet.iter_rows(values_only=True))

        self.assertEqual(rows[0], ('Title', 'Status'))
        self.assertEqual(
            rows[1:],
            [(f'row-{index}', 'pending') for index in range(6)],
        )

    @patch(
        'apps.tabdata.services.view_grid_service.get_grid_data_native',
        side_effect=TypeError('force ORM fallback'),
    )
    @patch('apps.tabdata.services.export_service._VIEW_EXPORT_PAGE_SIZE', 2)
    def test_excel_flattens_records_in_grouped_view_order(
        self,
        _native_query,
    ):
        import io

        import openpyxl

        result = ExportService(user=self.user).export_to_excel(
            table_id=self.table.id,
            view_id=self.view.id,
            field_ids=[self.title_field.id, self.status_field.id],
            include_headers=True,
            view_query={
                'filters': [],
                'sorts': [
                    {
                        'field_id': str(self.title_field.id),
                        'direction': 'desc',
                    }
                ],
                'groups': [
                    {
                        'field_id': str(self.status_field.id),
                        'direction': 'desc',
                    }
                ],
            },
        )

        worksheet = openpyxl.load_workbook(io.BytesIO(result), read_only=True).active
        rows = list(worksheet.iter_rows(values_only=True))

        self.assertEqual(rows[0], ('Title', 'Status'))
        self.assertEqual(
            rows[1:],
            [('row-7', 'fixed'), ('row-6', 'processing')]
            + [(f'row-{index}', 'pending') for index in reversed(range(6))]
            + [('row-8', None)],
        )

    @patch('apps.tabdata.services.export_service._VIEW_EXPORT_PAGE_SIZE', 2)
    def test_export_stats_reports_six_pending_rows_across_pages(self):
        stats = ExportService(user=self.user).get_export_stats(
            table_id=self.table.id,
            view_id=self.view.id,
            view_query=self._pending_view_query(),
        )

        self.assertEqual(stats['record_count'], 6)

    @patch('apps.tabdata.services.export_service._VIEW_EXPORT_PAGE_SIZE', 2)
    def test_csv_contains_six_pending_rows_across_pages(self):
        import csv
        import io

        content = ''.join(
            ExportService(user=self.user).export_to_csv_streaming(
                table_id=self.table.id,
                view_id=self.view.id,
                field_ids=[self.title_field.id, self.status_field.id],
                include_headers=True,
                view_query=self._pending_view_query(),
            )
        )
        rows = list(csv.reader(io.StringIO(content.lstrip('\ufeff'))))

        self.assertEqual(rows[0], ['Title', 'Status'])
        self.assertEqual(
            rows[1:],
            [[f'row-{index}', 'pending'] for index in range(6)],
        )

    @patch('apps.tabdata.services.export_service._VIEW_EXPORT_PAGE_SIZE', 2)
    def test_pdf_contains_six_pending_rows_across_pages(self):
        import io

        import pdfplumber

        content = ExportService(user=self.user).export_to_pdf(
            table_id=self.table.id,
            view_id=self.view.id,
            field_ids=[self.title_field.id, self.status_field.id],
            view_query=self._pending_view_query(),
        )
        with pdfplumber.open(io.BytesIO(content)) as document:
            text = '\n'.join(page.extract_text() or '' for page in document.pages)

        for index in range(6):
            self.assertIn(f'row-{index}', text)
        self.assertNotIn('row-6', text)
        self.assertNotIn('row-7', text)

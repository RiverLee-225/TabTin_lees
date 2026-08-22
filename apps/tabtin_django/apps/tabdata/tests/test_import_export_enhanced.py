"""
导入导出功能增强测试

测试CSV、Excel、JSON导入导出，数据预览，增量导入等功能
"""
import json
import io
import base64
import csv
import zipfile
from unittest.mock import patch
from uuid import uuid4
from django.test import TestCase, override_settings
from django.contrib.auth import get_user_model
from apps.tabtinspace.models import (
    Organization,
    OrganizationMember,
    OrganizationMemberIdentitySnapshot,
    Project,
    Space,
)
from apps.tabdata.models import AttachmentReference, Table, TableField, TableRecord, TableView
from apps.tabdata.services import ImportService, ExportService
from apps.tabdata.utils.record_data_access import read_data
from apps.services.oss.models import FileRecord

User = get_user_model()


class ImportExportEnhancedTestCase(TestCase):
    """导入导出增强功能测试"""

    databases = {'default', 'postgresql'}

    @staticmethod
    def _tiny_png_bytes():
        from PIL import Image

        image = Image.new('RGB', (1, 1), 'red')
        buffer = io.BytesIO()
        image.save(buffer, format='PNG')
        return buffer.getvalue()

    def setUp(self):
        """测试前准备"""
        # 创建测试用户
        self.user = User.objects.create_user(
            username='testuser',
            email='test@example.com',
            password='password123'
        )

        # 创建组织
        self.organization = Organization.objects.create(
            name='Test Organization',
            owner=self.user
        )

        # 添加成员
        OrganizationMember.objects.create(
            organization=self.organization,
            user=self.user,
            role='owner'
        )

        # 创建项目
        self.space = Space.objects.create(
            organization=self.organization,
            name='Test Space',
            description='Test project for testing',
        )

        # 创建表格
        self.table = Table.objects.create(
            space_id=self.space.id,
            organization_id=self.space.organization_id,
            name='Test Table',
            description='Test table',
            owner=self.user
        )

        # 创建字段
        self.field_name = TableField.objects.create(
            table=self.table,
            name='姓名',
            field_type='text',
            is_primary=True,
            order=0
        )

        self.field_age = TableField.objects.create(
            table=self.table,
            name='年龄',
            field_type='number',
            order=1
        )

        self.field_email = TableField.objects.create(
            table=self.table,
            name='邮箱',
            field_type='email',
            order=2
        )

    def test_csv_import_basic(self):
        """测试基础CSV导入"""
        csv_content = """姓名,年龄,邮箱
张三,25,zhang@example.com
李四,30,li@example.com
王五,28,wang@example.com"""

        service = ImportService(user=self.user)
        created, updated, errors = service.import_from_csv(
            table_id=self.table.id,
            file_content=csv_content,
            skip_errors=False
        )

        self.assertEqual(created, 3)
        self.assertEqual(updated, 0)
        self.assertEqual(len(errors), 0)

        # 验证记录
        records = TableRecord.objects.filter(table=self.table, is_deleted=False)
        self.assertEqual(records.count(), 3)

    def test_csv_import_incremental(self):
        """测试增量CSV导入（更新已有记录）"""
        # 先导入初始数据
        csv_content_1 = """姓名,年龄,邮箱
张三,25,zhang@example.com
李四,30,li@example.com"""

        service = ImportService(user=self.user)
        created, updated, errors = service.import_from_csv(
            table_id=self.table.id,
            file_content=csv_content_1,
            skip_errors=False
        )

        self.assertEqual(created, 2)

        # 增量导入，更新张三的年龄和邮箱，新增王五
        csv_content_2 = """姓名,年龄,邮箱
张三,26,zhang_new@example.com
王五,28,wang@example.com"""

        created, updated, errors = service.import_from_csv(
            table_id=self.table.id,
            file_content=csv_content_2,
            skip_errors=False,
            update_existing=True,
            primary_key_field='姓名'
        )

        self.assertEqual(created, 1)  # 王五是新增的
        self.assertEqual(updated, 1)  # 张三被更新了
        self.assertEqual(len(errors), 0)

        # 验证更新
        records = TableRecord.objects.filter(table=self.table, is_deleted=False)
        self.assertEqual(records.count(), 3)

        # 验证张三的数据被更新
        name_key = str(self.field_name.id)
        age_key = str(self.field_age.id)
        email_key = str(self.field_email.id)
        zhang_record = records.filter(data__contains={name_key: '张三'}).first()
        self.assertIsNotNone(zhang_record)
        self.assertEqual(read_data(zhang_record)[age_key], 26)
        self.assertEqual(read_data(zhang_record)[email_key], 'zhang_new@example.com')

    def test_csv_import_incremental_number_pk(self):
        """number 主键：CSV 字符串 / 结构化 int 都应匹配已存 float，走更新而非新建。"""
        number_table = Table.objects.create(
            space_id=self.space.id,
            organization_id=self.space.organization_id,
            name='Number PK Table',
            owner=self.user,
        )
        field_sku = TableField.objects.create(
            table=number_table,
            name='sku',
            field_type='number',
            is_primary=True,
            order=0,
        )
        field_label = TableField.objects.create(
            table=number_table,
            name='label',
            field_type='text',
            order=1,
        )
        service = ImportService(user=self.user)

        # 首次导入：CSV 单元格是字符串 "10"，入库经 NumberField.format 成 float 10.0
        created, updated, errors = service.import_from_csv(
            table_id=number_table.id,
            file_content='sku,label\n10,alpha\n11,beta\n0,zero\n',
            skip_errors=False,
        )
        self.assertEqual(errors, [])
        self.assertEqual(created, 3)
        self.assertEqual(updated, 0)

        # 增量：同 sku，改 label；应全部 update，含合法主键 0
        created, updated, errors = service.import_from_csv(
            table_id=number_table.id,
            file_content='sku,label\n10,alpha2\n11,beta2\n0,zero2\n',
            skip_errors=False,
            update_existing=True,
            primary_key_field='sku',
        )
        self.assertEqual(errors, [])
        self.assertEqual(created, 0)
        self.assertEqual(updated, 3)
        self.assertEqual(
            TableRecord.objects.filter(table=number_table, is_deleted=False).count(),
            3,
        )

        sku_key = str(field_sku.id)
        label_key = str(field_label.id)
        records_by_sku = {
            read_data(r)[sku_key]: read_data(r)[label_key]
            for r in TableRecord.objects.filter(table=number_table, is_deleted=False)
        }
        # 存库可能是 int 或 float，按数值对齐
        normalized = {float(k): v for k, v in records_by_sku.items()}
        self.assertEqual(normalized[10.0], 'alpha2')
        self.assertEqual(normalized[11.0], 'beta2')
        self.assertEqual(normalized[0.0], 'zero2')

        # 模拟 Excel 读出的 int 单元格（结构化 JSON），再增量更新
        json_content = json.dumps({
            'headers': ['sku', 'label'],
            'data': [
                [10, 'alpha3'],
                [11, 'beta3'],
                [0, 'zero3'],
                [12, 'gamma'],  # 新行
            ],
        })
        created, updated, errors = service.import_from_json(
            table_id=number_table.id,
            json_content=json_content,
            skip_errors=False,
            update_existing=True,
            primary_key_field=str(field_sku.id),
        )
        self.assertEqual(errors, [])
        self.assertEqual(created, 1)
        self.assertEqual(updated, 3)
        self.assertEqual(
            TableRecord.objects.filter(table=number_table, is_deleted=False).count(),
            4,
        )

    def test_json_import_array_format(self):
        """测试JSON导入（数组格式）"""
        json_content = json.dumps([
            {"姓名": "张三", "年龄": 25, "邮箱": "zhang@example.com"},
            {"姓名": "李四", "年龄": 30, "邮箱": "li@example.com"},
        ], ensure_ascii=False)

        service = ImportService(user=self.user)
        created, updated, errors = service.import_from_json(
            table_id=self.table.id,
            json_content=json_content,
            skip_errors=False
        )

        self.assertEqual(created, 2)
        self.assertEqual(updated, 0)
        self.assertEqual(len(errors), 0)

    def test_json_import_structured_format(self):
        """测试JSON导入（结构化格式）"""
        json_content = json.dumps({
            "headers": ["姓名", "年龄", "邮箱"],
            "data": [
                ["张三", 25, "zhang@example.com"],
                ["李四", 30, "li@example.com"]
            ]
        }, ensure_ascii=False)

        service = ImportService(user=self.user)
        created, updated, errors = service.import_from_json(
            table_id=self.table.id,
            json_content=json_content,
            skip_errors=False
        )

        self.assertEqual(created, 2)
        self.assertEqual(updated, 0)

    def test_json_import_table_full_format(self):
        """测试 JSON 导入（table_full 快照格式）"""
        source_name_id = str(uuid4())
        source_age_id = str(uuid4())
        source_email_id = str(uuid4())

        json_content = json.dumps({
            "id": str(uuid4()),
            "name": "Source Table",
            "fields": [
                {"id": source_name_id, "name": "姓名", "field_type": "text", "order": 0},
                {"id": source_age_id, "name": "年龄", "field_type": "number", "order": 1},
                {"id": source_email_id, "name": "邮箱", "field_type": "email", "order": 2},
            ],
            "views": [],
            "records": [
                {
                    "id": str(uuid4()),
                    "fields": {
                        source_name_id: "张三",
                        source_age_id: 25,
                        source_email_id: "zhang@example.com",
                    },
                },
                {
                    "id": str(uuid4()),
                    "fields": {
                        source_name_id: "李四",
                        source_age_id: 30,
                        source_email_id: "li@example.com",
                    },
                },
            ],
            "metadata": {
                "format": "table_full",
                "format_version": "tabtin.table_full.v1",
            },
        }, ensure_ascii=False)

        service = ImportService(user=self.user)
        created, updated, errors = service.import_from_json(
            table_id=self.table.id,
            json_content=json_content,
            skip_errors=False
        )

        self.assertEqual(created, 2)
        self.assertEqual(updated, 0)
        self.assertEqual(errors, [])

        records = TableRecord.objects.filter(table=self.table, is_deleted=False).order_by('order')
        self.assertEqual(records.count(), 2)

        name_key = str(self.field_name.id)
        age_key = str(self.field_age.id)
        email_key = str(self.field_email.id)

        first = records.first()
        self.assertIsNotNone(first)
        self.assertEqual(read_data(first)[name_key], "张三")
        self.assertEqual(read_data(first)[age_key], 25)
        self.assertEqual(read_data(first)[email_key], "zhang@example.com")

    def test_json_import_table_full_format_supports_data_key_and_name_key(self):
        """测试 table_full 导入兼容 records.data + 字段名 key"""
        source_name_id = str(uuid4())
        source_age_id = str(uuid4())
        source_email_id = str(uuid4())

        json_content = json.dumps({
            "fields": [
                {"id": source_email_id, "name": "邮箱", "order": 2},
                {"id": source_name_id, "name": "姓名", "order": 0},
                {"id": source_age_id, "name": "年龄", "order": 1},
            ],
            "records": [
                {
                    "id": str(uuid4()),
                    "data": {
                        "姓名": "王五",
                        source_age_id: 27,
                        "邮箱": "wang@example.com",
                    },
                },
            ],
            "metadata": {"format": "table_full"},
        }, ensure_ascii=False)

        service = ImportService(user=self.user)
        created, updated, errors = service.import_from_json(
            table_id=self.table.id,
            json_content=json_content,
            skip_errors=False
        )

        self.assertEqual(created, 1)
        self.assertEqual(updated, 0)
        self.assertEqual(errors, [])

        record = TableRecord.objects.filter(table=self.table, is_deleted=False).latest('created_at')
        self.assertEqual(read_data(record)[str(self.field_name.id)], "王五")
        self.assertEqual(read_data(record)[str(self.field_age.id)], 27)
        self.assertEqual(read_data(record)[str(self.field_email.id)], "wang@example.com")

    def test_json_import_table_full_format_invalid_shape(self):
        """测试 table_full 非法结构返回可诊断错误"""
        invalid_json = json.dumps({
            "fields": {"bad": "shape"},
            "records": [],
            "metadata": {"format": "table_full"},
        }, ensure_ascii=False)

        service = ImportService(user=self.user)
        created, updated, errors = service.import_from_json(
            table_id=self.table.id,
            json_content=invalid_json,
            skip_errors=False
        )

        self.assertEqual(created, 0)
        self.assertEqual(updated, 0)
        self.assertTrue(errors)
        self.assertIn("table_full 格式非法", errors[0])

    def test_json_import_table_full_format_allows_empty_records(self):
        """测试 table_full 导入允许空记录（仅导入结构/视图）"""
        source_name_id = str(uuid4())
        source_view_id = str(uuid4())

        json_content = json.dumps({
            "fields": [
                {"id": source_name_id, "name": "姓名", "field_type": "text", "order": 0},
            ],
            "records": [],
            "views": [
                {
                    "id": source_view_id,
                    "name": "空表视图",
                    "view_type": "grid",
                    "visible_fields": [source_name_id],
                    "field_order": [source_name_id],
                }
            ],
            "metadata": {"format": "table_full"},
        }, ensure_ascii=False)

        service = ImportService(user=self.user)
        created, updated, errors = service.import_from_json(
            table_id=self.table.id,
            json_content=json_content,
            skip_errors=False,
        )

        self.assertEqual(created, 0)
        self.assertEqual(updated, 0)
        self.assertEqual(errors, [])
        self.assertTrue(TableView.objects.filter(table=self.table, name="空表视图").exists())

    def test_json_import_table_full_format_imports_views_and_remaps_field_refs(self):
        """测试 table_full 导入会创建视图并重写字段引用"""
        source_name_id = str(uuid4())
        source_age_id = str(uuid4())
        source_email_id = str(uuid4())

        json_content = json.dumps({
            "fields": [
                {"id": source_name_id, "name": "姓名", "order": 0},
                {"id": source_age_id, "name": "年龄", "order": 1},
                {"id": source_email_id, "name": "邮箱", "order": 2},
            ],
            "records": [
                {
                    "id": str(uuid4()),
                    "fields": {
                        source_name_id: "赵六",
                        source_age_id: 19,
                        source_email_id: "zhao@example.com",
                    },
                }
            ],
            "views": [
                {
                    "name": "导入视图",
                    "view_type": "grid",
                    "visible_fields": [source_name_id, source_email_id],
                    "field_order": [source_name_id, source_age_id, source_email_id],
                    "column_meta": {
                        source_name_id: {"order": 0, "width": 200},
                        source_age_id: {"order": 1, "hidden": True},
                        source_email_id: {"order": 2},
                    },
                    "filters": [
                        {"field_id": source_age_id, "operator": "greater_than", "value": 18},
                    ],
                    "sorts": [
                        {"field": source_name_id, "direction": "asc"},
                    ],
                    "groups": [
                        {"field": source_email_id},
                    ],
                    "config": {
                        "column_widths": {
                            source_name_id: 200,
                            source_age_id: 120,
                        }
                    },
                }
            ],
            "metadata": {"format": "table_full"},
        }, ensure_ascii=False)

        service = ImportService(user=self.user)
        created, updated, errors = service.import_from_json(
            table_id=self.table.id,
            json_content=json_content,
            skip_errors=False
        )

        self.assertEqual(created, 1)
        self.assertEqual(updated, 0)
        self.assertEqual(errors, [])

        view = TableView.objects.get(table=self.table, name='导入视图')
        target_name_id = str(self.field_name.id)
        target_age_id = str(self.field_age.id)
        target_email_id = str(self.field_email.id)

        self.assertEqual(view.visible_fields, [target_name_id, target_email_id])
        self.assertEqual(view.field_order, [target_name_id, target_age_id, target_email_id])
        self.assertIn(target_name_id, view.column_meta)
        self.assertIn(target_age_id, view.column_meta)
        self.assertNotIn(source_name_id, view.column_meta)
        self.assertEqual(view.column_meta[target_name_id]['width'], 200)
        self.assertEqual(view.filters[0]['field_id'], target_age_id)
        self.assertEqual(view.sorts[0]['field'], target_name_id)
        self.assertEqual(view.groups[0]['field'], target_email_id)
        self.assertEqual(view.config['column_widths'][target_name_id], 200)

    def test_field_type_inference(self):
        """测试字段类型推断"""
        service = ImportService(user=self.user)

        # 测试数字推断
        self.assertEqual(service.infer_field_type(['123', '456', '789']), 'number')

        # 测试日期推断
        self.assertEqual(service.infer_field_type(['2025-01-01', '2025-02-01', '2025-03-01']), 'date')
        self.assertEqual(service.infer_field_type(['2025/01/01', '2025/02/01', '2025/03/01']), 'date')
        self.assertEqual(service.infer_field_type(['2025年01月01日', '2025年02月01日', '2025年03月01日']), 'date')
        self.assertEqual(
            service.infer_field_type(['2025-01-01T10:30:00Z', '2025-01-02T11:45:00Z']),
            'text',
        )

        # UUID / 关联引用类文本不能只因为包含横杠就被误判成日期
        self.assertEqual(
            service.infer_field_type([
                'b1f3e884-6c2b-4e63-bbc4-daed4dd74f8d',
                'cf539271-fa15-4c3a-be43-61189ba9fa61',
                '货架C-父记录',
            ]),
            'text'
        )

        # 测试布尔值推断
        self.assertEqual(service.infer_field_type(['true', 'false', 'true']), 'checkbox')

        # 测试选项推断（重复值多）
        self.assertEqual(
            service.infer_field_type(['选项1', '选项2', '选项1', '选项2', '选项1', '选项2', '选项1']),
            'select'
        )

        # 测试文本（默认）
        self.assertEqual(service.infer_field_type(['随机文本1', '随机文本2', '随机文本3']), 'text')
        # URL 推断回归见 test_import_url_inference_6144.py（，SimpleTestCase）

    def test_smart_field_mapping(self):
        """测试智能字段匹配"""
        headers = ['姓名', '年龄', '邮箱', '地址']  # 地址是新字段
        rows = [
            ['张三', '25', 'zhang@example.com', '北京市'],
            ['李四', '30', 'li@example.com', '上海市']
        ]

        service = ImportService(user=self.user)
        result = service.smart_field_mapping(self.table.id, headers, rows)

        # 验证字段映射
        self.assertIn('姓名', result['field_mapping'])
        self.assertIn('年龄', result['field_mapping'])
        self.assertIn('邮箱', result['field_mapping'])

        # 验证新字段建议
        self.assertEqual(len(result['new_fields']), 1)
        self.assertEqual(result['new_fields'][0]['name'], '地址')
        self.assertEqual(result['new_fields'][0]['type'], 'text')

    def test_csv_export_basic(self):
        """测试基础CSV导出"""
        # 先创建一些记录
        name_key = str(self.field_name.id)
        age_key = str(self.field_age.id)
        email_key = str(self.field_email.id)
        TableRecord.objects.bulk_create([
            TableRecord(
                table=self.table,
                data={name_key: '张三', age_key: 25, email_key: 'zhang@example.com'},
                created_by=self.user
            ),
            TableRecord(
                table=self.table,
                data={name_key: '李四', age_key: 30, email_key: 'li@example.com'},
                created_by=self.user
            )
        ])

        service = ExportService(user=self.user)
        csv_content = service.export_to_csv(
            table_id=self.table.id,
            include_headers=True
        )

        self.assertIsNotNone(csv_content)
        self.assertIn('姓名', csv_content)
        self.assertIn('张三', csv_content)
        self.assertIn('李四', csv_content)

    def test_csv_export_reads_collab_hex_keys(self):
        """CSV 导出应兼容协作 persist 写入的 field.id.hex key。"""
        TableRecord.objects.create(
            table=self.table,
            data={
                self.field_name.id.hex: '协作标题',
                self.field_age.id.hex: 31,
                self.field_email.id.hex: 'collab@example.com',
            },
            created_by=self.user,
        )

        service = ExportService(user=self.user)
        csv_content = service.export_to_csv(
            table_id=self.table.id,
            include_headers=True,
        )

        self.assertIn('协作标题', csv_content)
        self.assertIn('31', csv_content)
        self.assertIn('collab@example.com', csv_content)

    def test_csv_export_preserves_requested_record_id_order(self):
        """按选中记录导出时，应保留前端传入的记录顺序。"""
        name_key = str(self.field_name.id)
        first = TableRecord.objects.create(
            table=self.table,
            data={name_key: '第一行'},
            created_by=self.user,
            order=10,
        )
        second = TableRecord.objects.create(
            table=self.table,
            data={name_key: '第二行'},
            created_by=self.user,
            order=20,
        )
        third = TableRecord.objects.create(
            table=self.table,
            data={name_key: '第三行'},
            created_by=self.user,
            order=30,
        )

        service = ExportService(user=self.user)
        csv_content = service.export_to_csv(
            table_id=self.table.id,
            record_ids=[third.id, first.id, second.id],
            include_headers=True,
        )
        rows = list(csv.reader(io.StringIO(csv_content.lstrip('\ufeff'))))
        name_col_idx = rows[0].index('姓名')

        self.assertEqual(
            [row[name_col_idx] for row in rows[1:]],
            ['第三行', '第一行', '第二行'],
        )

    def test_export_stats_uses_current_view_visible_fields(self):
        """当前视图导出统计应与实际导出的可见字段一致。"""
        view = TableView.objects.create(
            table=self.table,
            name='只显示两列',
            view_type='grid',
            column_meta={
                str(self.field_name.id): {'order': 0, 'hidden': False},
                str(self.field_age.id): {'order': 1, 'hidden': False},
                str(self.field_email.id): {'order': 2, 'hidden': True},
            },
            created_by=self.user,
        )

        service = ExportService(user=self.user)
        stats = service.get_export_stats(table_id=self.table.id, view_id=view.id)

        self.assertEqual(stats['field_count'], 2)

    def test_csv_export_preserves_requested_field_id_order(self):
        """手动选择字段导出时，应保留前端传入的字段顺序。"""
        service = ExportService(user=self.user)
        csv_content = service.export_to_csv(
            table_id=self.table.id,
            field_ids=[self.field_email.id, self.field_name.id],
            include_headers=True,
        )
        rows = list(csv.reader(io.StringIO(csv_content.lstrip('\ufeff'))))

        self.assertEqual(rows[0], ['邮箱', '姓名'])

    def test_export_with_view_from_other_table_is_rejected(self):
        """显式 view_id 不属于当前表时，不应静默回退成全表导出。"""
        other_table = Table.objects.create(
            space_id=self.space.id,
            organization_id=self.space.organization_id,
            name='Other Table',
            owner=self.user,
        )
        other_view = TableView.objects.create(
            table=other_table,
            name='Other View',
            view_type='grid',
            created_by=self.user,
        )

        service = ExportService(user=self.user)
        with self.assertRaises(ValueError):
            service.export_to_csv(
                table_id=self.table.id,
                view_id=other_view.id,
                include_headers=True,
            )

    def test_export_stats_with_view_from_other_table_is_rejected(self):
        """导出统计也应拒绝跨表 view_id，避免接口静默统计全表。"""
        other_table = Table.objects.create(
            space_id=self.space.id,
            organization_id=self.space.organization_id,
            name='Other Table For Stats',
            owner=self.user,
        )
        other_view = TableView.objects.create(
            table=other_table,
            name='Other View For Stats',
            view_type='grid',
            created_by=self.user,
        )

        service = ExportService(user=self.user)
        with self.assertRaises(ValueError):
            service.get_export_stats(table_id=self.table.id, view_id=other_view.id)

    @patch('apps.tabdata.services.export_service.read_data_bulk')
    def test_csv_export_uses_native_preloaded_data(self, mock_read_data_bulk):
        """native-only 记录 JSONField 为空时，导出应读取预加载的 native 数据。"""
        TableRecord.objects.create(
            table=self.table,
            data={},
            created_by=self.user,
        )

        def preload(records, table, fields):
            for record in records:
                object.__setattr__(record, '_rda_cached_data', {
                    str(self.field_name.id): 'native 标题',
                    str(self.field_age.id): 42,
                    str(self.field_email.id): 'native@example.com',
                })

        mock_read_data_bulk.side_effect = preload

        service = ExportService(user=self.user)
        csv_content = service.export_to_csv(
            table_id=self.table.id,
            include_headers=True,
        )

        self.assertTrue(mock_read_data_bulk.called)
        self.assertIn('native 标题', csv_content)
        self.assertIn('42', csv_content)
        self.assertIn('native@example.com', csv_content)

    def test_csv_export_formats_single_link_dict(self):
        """父记录等单值 link 应导出可读标题，而不是原始 JSON。"""
        parent_field = TableField.objects.create(
            table=self.table,
            name='父记录',
            field_type='link',
            order=3,
            config={'foreignTableId': str(self.table.id), 'relationship': 'ManyOne'},
        )
        TableRecord.objects.create(
            table=self.table,
            data={str(parent_field.id): {'id': 'parent-1', 'title': '父记录标题'}},
            created_by=self.user,
        )

        service = ExportService(user=self.user)
        csv_content = service.export_to_csv(
            table_id=self.table.id,
            include_headers=True,
        )

        self.assertIn('父记录标题', csv_content)
        self.assertNotIn('{"id":', csv_content)

    def test_csv_export_preserves_image_and_non_image_attachment_names(self):
        """CSV 导出：图片和非图片附件都保留可见文件名。"""
        image_field = TableField.objects.create(
            table=self.table,
            name='图片',
            field_type='attachment',
            order=3,
        )
        image_bytes = self._tiny_png_bytes()
        TableRecord.objects.create(
            table=self.table,
            data={
                str(self.field_name.id): '带图记录',
                str(image_field.id): [
                    {
                        'name': 'inline.png',
                        'url': 'data:image/png;base64,' + base64.b64encode(image_bytes).decode('ascii'),
                        'size': len(image_bytes),
                        'mime_type': 'image/png',
                    },
                    {
                        'name': 'notes.pdf',
                        'url': 'https://assets.example.test/notes.pdf',
                        'size': 1024,
                        'mime_type': 'application/pdf',
                    },
                ],
            },
            created_by=self.user,
        )

        service = ExportService(user=self.user)
        csv_content = service.export_to_csv(table_id=self.table.id, include_headers=True)
        rows = list(csv.reader(io.StringIO(csv_content.lstrip('\ufeff'))))
        image_col_idx = rows[0].index('图片')

        self.assertEqual(rows[1][image_col_idx], 'inline.png, notes.pdf')
        self.assertIn('inline.png', csv_content)

    def test_csv_streaming_export_preserves_image_and_non_image_attachment_names(self):
        """流式 CSV 导出：图片和非图片附件都保留可见文件名。"""
        image_field = TableField.objects.create(
            table=self.table,
            name='图片',
            field_type='attachment',
            order=3,
        )
        image_bytes = self._tiny_png_bytes()
        TableRecord.objects.create(
            table=self.table,
            data={
                str(self.field_name.id): '带图记录',
                str(image_field.id): [
                    {
                        'name': 'inline.png',
                        'url': 'data:image/png;base64,' + base64.b64encode(image_bytes).decode('ascii'),
                        'size': len(image_bytes),
                        'mime_type': 'image/png',
                    },
                    {
                        'name': 'notes.pdf',
                        'url': 'https://assets.example.test/notes.pdf',
                        'size': 1024,
                        'mime_type': 'application/pdf',
                    },
                ],
            },
            created_by=self.user,
        )

        service = ExportService(user=self.user)
        streaming_content = ''.join(
            service.export_to_csv_streaming(table_id=self.table.id, include_headers=True)
        )
        streaming_rows = list(csv.reader(io.StringIO(streaming_content.lstrip('\ufeff'))))
        streaming_image_col_idx = streaming_rows[0].index('图片')

        self.assertEqual(streaming_rows[1][streaming_image_col_idx], 'inline.png, notes.pdf')
        self.assertIn('inline.png', streaming_content)

    def test_json_export_array_format(self):
        """测试JSON导出（数组格式）"""
        # 创建记录
        name_key = str(self.field_name.id)
        age_key = str(self.field_age.id)
        email_key = str(self.field_email.id)
        TableRecord.objects.create(
            table=self.table,
            data={name_key: '张三', age_key: 25, email_key: 'zhang@example.com'},
            created_by=self.user
        )

        service = ExportService(user=self.user)
        json_content = service.export_to_json(
            table_id=self.table.id,
            format_type='array'
        )

        self.assertIsNotNone(json_content)
        data = json.loads(json_content)
        self.assertIsInstance(data, list)
        self.assertEqual(len(data), 1)
        self.assertEqual(data[0]['姓名'], '张三')

    def test_json_export_preserves_image_attachment_metadata(self):
        """JSON 导出面向程序/API，保留图片附件原始结构。"""
        image_field = TableField.objects.create(
            table=self.table,
            name='图片',
            field_type='attachment',
            order=3,
        )
        image_bytes = self._tiny_png_bytes()
        TableRecord.objects.create(
            table=self.table,
            data={
                str(self.field_name.id): '带图记录',
                str(image_field.id): [
                    {
                        'name': 'inline.png',
                        'url': 'data:image/png;base64,' + base64.b64encode(image_bytes).decode('ascii'),
                        'size': len(image_bytes),
                        'mime_type': 'image/png',
                    }
                ],
            },
            created_by=self.user,
        )

        service = ExportService(user=self.user)
        json_content = service.export_to_json(table_id=self.table.id, format_type='array')
        data = json.loads(json_content)

        self.assertEqual(data[0]['图片'][0]['name'], 'inline.png')

    def test_json_streaming_export_preserves_image_attachment_metadata(self):
        """流式 JSON array 导出同样保留图片附件原始结构。"""
        image_field = TableField.objects.create(
            table=self.table,
            name='图片',
            field_type='attachment',
            order=3,
        )
        image_bytes = self._tiny_png_bytes()
        TableRecord.objects.create(
            table=self.table,
            data={
                str(self.field_name.id): '带图记录',
                str(image_field.id): [
                    {
                        'name': 'inline.png',
                        'url': 'data:image/png;base64,' + base64.b64encode(image_bytes).decode('ascii'),
                        'size': len(image_bytes),
                        'mime_type': 'image/png',
                    }
                ],
            },
            created_by=self.user,
        )

        service = ExportService(user=self.user)
        json_content = ''.join(service.export_to_json_streaming(table_id=self.table.id))
        data = json.loads(json_content)

        self.assertEqual(data[0]['图片'][0]['name'], 'inline.png')

    def test_json_export_structured_format(self):
        """测试JSON导出（结构化格式）"""
        # 创建记录
        name_key = str(self.field_name.id)
        age_key = str(self.field_age.id)
        email_key = str(self.field_email.id)
        TableRecord.objects.create(
            table=self.table,
            data={name_key: '张三', age_key: 25, email_key: 'zhang@example.com'},
            created_by=self.user
        )

        service = ExportService(user=self.user)
        json_content = service.export_to_json(
            table_id=self.table.id,
            format_type='structured'
        )

        self.assertIsNotNone(json_content)
        data = json.loads(json_content)
        self.assertIn('headers', data)
        self.assertIn('data', data)
        self.assertIn('metadata', data)
        self.assertEqual(data['headers'], ['姓名', '年龄', '邮箱'])

    def test_json_export_table_full_format(self):
        """测试 JSON 导出（table_full 格式）"""
        name_key = str(self.field_name.id)
        age_key = str(self.field_age.id)
        email_key = str(self.field_email.id)

        record = TableRecord.objects.create(
            table=self.table,
            data={name_key: '张三', age_key: 25, email_key: 'zhang@example.com'},
            created_by=self.user
        )

        view = TableView.objects.create(
            table=self.table,
            name='主视图',
            view_type='grid',
            visible_fields=[name_key, email_key],
            field_order=[name_key, age_key, email_key],
            column_meta={
                name_key: {'order': 0, 'width': 220},
                age_key: {'order': 1, 'hidden': True},
                email_key: {'order': 2},
            },
            created_by=self.user,
            order=0
        )

        service = ExportService(user=self.user)
        json_content = service.export_to_json(
            table_id=self.table.id,
            format_type='table_full'
        )

        self.assertIsNotNone(json_content)
        data = json.loads(json_content)

        self.assertEqual(data['id'], str(self.table.id))
        self.assertIn('fields', data)
        self.assertIn('views', data)
        self.assertIn('records', data)
        self.assertEqual(data['metadata']['format'], 'table_full')
        self.assertEqual(data['metadata']['format_version'], 'tabtin.table_full.v1')

        self.assertEqual(len(data['fields']), 3)
        exported_view = next(item for item in data['views'] if item['id'] == str(view.id))
        self.assertIn('column_meta', exported_view)
        self.assertIn('columnMeta', exported_view)
        self.assertEqual(exported_view['column_meta'], exported_view['columnMeta'])
        self.assertEqual(exported_view['type'], exported_view['view_type'])
        self.assertEqual(exported_view['filter'], exported_view['filters'])
        self.assertEqual(exported_view['sort'], exported_view['sorts'])
        self.assertEqual(exported_view['group'], exported_view['groups'])
        self.assertEqual(exported_view['options'], exported_view['config'])

        exported_record = next(item for item in data['records'] if item['id'] == str(record.id))
        self.assertIn('fields', exported_record)
        self.assertIn('data', exported_record)
        self.assertEqual(exported_record['fields'], exported_record['data'])
        self.assertEqual(exported_record['fields'][name_key], '张三')

        first_field = data['fields'][0]
        self.assertIn('type', first_field)
        self.assertIn('options', first_field)
        self.assertIn('isPrimary', first_field)
        self.assertIn('notNull', first_field)
        self.assertEqual(first_field['options'], first_field['config'])

    def test_json_export_project_base_full_format(self):
        """测试项目级 JSON 导出（base_full）"""
        name_key = str(self.field_name.id)
        age_key = str(self.field_age.id)
        email_key = str(self.field_email.id)
        TableRecord.objects.create(
            table=self.table,
            data={name_key: '张三', age_key: 25, email_key: 'zhang@example.com'},
            created_by=self.user
        )

        table2 = Table.objects.create(
            space_id=self.space.id,
            organization_id=self.space.organization_id,
            name='Test Table 2',
            description='Secondary table',
            owner=self.user
        )
        field2_title = TableField.objects.create(
            table=table2,
            name='标题',
            field_type='text',
            is_primary=True,
            order=0
        )
        TableRecord.objects.create(
            table=table2,
            data={str(field2_title.id): '第二张表记录'},
            created_by=self.user
        )

        service = ExportService(user=self.user)
        json_content = service.export_space_to_json(
            space_id=self.space.id,
            format_type='base_full'
        )

        self.assertIsNotNone(json_content)
        data = json.loads(json_content)
        self.assertEqual(data['id'], str(self.space.id))
        self.assertIn('tables', data)
        self.assertEqual(data['metadata']['format'], 'base_full')
        self.assertEqual(data['metadata']['format_version'], 'tabtin.base_full.v1')
        self.assertEqual(data['metadata']['table_count'], 2)

        table_ids = {item['id'] for item in data['tables']}
        self.assertIn(str(self.table.id), table_ids)
        self.assertIn(str(table2.id), table_ids)
        first_table = next(item for item in data['tables'] if item['id'] == str(self.table.id))
        self.assertEqual(first_table['metadata']['format'], 'table_full')
        self.assertIn('fields', first_table)
        self.assertIn('views', first_table)
        self.assertIn('records', first_table)

    def test_json_export_project_base_full_with_table_ids_filter(self):
        """测试项目级 JSON 导出按 table_ids 过滤"""
        table2 = Table.objects.create(
            space_id=self.space.id,
            organization_id=self.space.organization_id,
            name='Filtered Out Table',
            description='Should be filtered',
            owner=self.user
        )
        TableField.objects.create(
            table=table2,
            name='名称',
            field_type='text',
            is_primary=True,
            order=0
        )

        service = ExportService(user=self.user)
        json_content = service.export_space_to_json(
            space_id=self.space.id,
            table_ids=[self.table.id],
            format_type='base_full'
        )

        self.assertIsNotNone(json_content)
        data = json.loads(json_content)
        self.assertEqual(data['metadata']['table_count'], 1)
        self.assertEqual(len(data['tables']), 1)
        self.assertEqual(data['tables'][0]['id'], str(self.table.id))

    def test_json_import_project_base_full_format(self):
        """测试项目级 JSON 导入（base_full）"""
        source_name_id = str(uuid4())
        source_age_id = str(uuid4())
        source_email_id = str(uuid4())
        source_view_id = str(uuid4())

        payload = {
            "id": str(uuid4()),
            "name": "Source Space",
            "tables": [
                {
                    "id": str(uuid4()),
                    "name": "Imported Table",
                    "description": "from snapshot",
                    "fields": [
                        {"id": source_name_id, "name": "姓名", "field_type": "text", "is_primary": True, "order": 0},
                        {"id": source_age_id, "name": "年龄", "field_type": "number", "order": 1},
                        {"id": source_email_id, "name": "邮箱", "field_type": "email", "order": 2},
                    ],
                    "records": [
                        {
                            "id": str(uuid4()),
                            "fields": {
                                source_name_id: "张三",
                                source_age_id: 25,
                                source_email_id: "zhang@example.com",
                            },
                        },
                    ],
                    "views": [
                        {
                            "id": source_view_id,
                            "name": "主视图",
                            "view_type": "grid",
                            "visible_fields": [source_name_id, source_email_id],
                            "field_order": [source_name_id, source_age_id, source_email_id],
                            "filters": [
                                {"field_id": source_age_id, "operator": "greater_than", "value": 18},
                            ],
                            "config": {
                                "column_widths": {
                                    source_name_id: 220,
                                }
                            },
                        }
                    ],
                    "default_view_id": source_view_id,
                    "metadata": {"format": "table_full"},
                }
            ],
            "metadata": {
                "format": "base_full",
                "format_version": "tabtin.base_full.v1",
            },
        }

        service = ImportService(user=self.user)
        result = service.import_space_from_json(
            space_id=self.space.id,
            json_content=json.dumps(payload, ensure_ascii=False),
            skip_errors=False,
        )

        self.assertEqual(result["created_tables"], 1)
        self.assertEqual(result["created_count"], 1)
        self.assertEqual(result["updated_count"], 0)
        self.assertEqual(result["errors"], [])
        self.assertEqual(len(result["table_results"]), 1)

        imported_table = Table.objects.get(space_id=self.space.id, name="Imported Table")
        self.assertEqual(imported_table.field_count, 3)
        self.assertEqual(imported_table.row_count, 1)

        fields = {
            field.name: field
            for field in TableField.objects.filter(table=imported_table, is_deleted=False)
        }
        record = TableRecord.objects.get(table=imported_table, is_deleted=False)
        self.assertEqual(read_data(record)[str(fields["姓名"].id)], "张三")
        self.assertEqual(read_data(record)[str(fields["年龄"].id)], 25)
        self.assertEqual(read_data(record)[str(fields["邮箱"].id)], "zhang@example.com")

        views = list(TableView.objects.filter(table=imported_table).order_by('order'))
        self.assertEqual(len(views), 1)
        imported_view = views[0]
        self.assertEqual(str(imported_table.default_view_id), str(imported_view.id))
        self.assertEqual(imported_view.name, "主视图")
        self.assertEqual(imported_view.filters[0]["field_id"], str(fields["年龄"].id))
        self.assertIn(str(fields["姓名"].id), imported_view.config["column_widths"])

    def test_json_import_project_base_full_resolves_table_name_conflict(self):
        """测试项目级导入遇到同名表时自动重命名"""
        payload = {
            "tables": [
                {
                    "name": self.table.name,
                    "fields": [
                        {"id": str(uuid4()), "name": "标题", "field_type": "text", "is_primary": True, "order": 0},
                    ],
                    "records": [
                        {"id": str(uuid4()), "fields": {"标题": "冲突后的新表"}},
                    ],
                    "views": [],
                }
            ],
            "metadata": {"format": "base_full"},
        }

        service = ImportService(user=self.user)
        result = service.import_space_from_json(
            space_id=self.space.id,
            json_content=json.dumps(payload, ensure_ascii=False),
            skip_errors=False,
        )

        self.assertEqual(result["created_tables"], 1)
        self.assertEqual(result["errors"], [])
        imported_name = result["table_results"][0]["table_name"]
        self.assertNotEqual(imported_name, self.table.name)
        self.assertTrue(imported_name.startswith(self.table.name))
        self.assertTrue(Table.objects.filter(space_id=self.space.id, name=imported_name).exists())

    def test_json_import_project_base_full_skip_errors_continues(self):
        """测试项目级导入在 skip_errors=True 下可跳过坏表继续导入"""
        payload = {
            "tables": [
                {
                    "name": "Bad Table",
                    "fields": {"bad": "shape"},
                    "records": [],
                },
                {
                    "name": "Good Table",
                    "fields": [
                        {"id": str(uuid4()), "name": "标题", "field_type": "text", "is_primary": True, "order": 0},
                    ],
                    "records": [
                        {"id": str(uuid4()), "fields": {"标题": "可导入记录"}},
                    ],
                    "views": [],
                }
            ],
            "metadata": {"format": "base_full"},
        }

        service = ImportService(user=self.user)
        result = service.import_space_from_json(
            space_id=self.space.id,
            json_content=json.dumps(payload, ensure_ascii=False),
            skip_errors=True,
        )

        self.assertEqual(result["created_tables"], 1)
        self.assertEqual(result["created_count"], 1)
        self.assertGreaterEqual(len(result["errors"]), 1)
        self.assertIn("第1张表导入失败", result["errors"][0])
        self.assertTrue(Table.objects.filter(space_id=self.space.id, name="Good Table").exists())

    def test_json_import_project_base_full_allows_empty_table(self):
        """测试项目级导入允许空记录表"""
        payload = {
            "tables": [
                {
                    "name": "Empty Table",
                    "fields": [
                        {"id": str(uuid4()), "name": "标题", "field_type": "text", "is_primary": True, "order": 0},
                    ],
                    "records": [],
                    "views": [],
                }
            ],
            "metadata": {"format": "base_full"},
        }

        service = ImportService(user=self.user)
        result = service.import_space_from_json(
            space_id=self.space.id,
            json_content=json.dumps(payload, ensure_ascii=False),
            skip_errors=False,
        )

        self.assertEqual(result["created_tables"], 1)
        self.assertEqual(result["created_count"], 0)
        self.assertEqual(result["updated_count"], 0)
        self.assertEqual(result["errors"], [])
        imported_table = Table.objects.get(space_id=self.space.id, name="Empty Table")
        self.assertEqual(imported_table.row_count, 0)
        self.assertEqual(imported_table.field_count, 1)
        self.assertIsNotNone(imported_table.default_view_id)

    def test_json_import_project_base_full_remaps_cross_table_field_config_refs(self):
        """测试项目级导入会重写跨表字段配置引用（table/field/公式）"""
        source_table_a_id = str(uuid4())
        source_table_b_id = str(uuid4())
        source_a_title_id = str(uuid4())
        source_a_link_id = str(uuid4())
        source_a_formula_id = str(uuid4())
        source_b_title_id = str(uuid4())
        source_b_view_id = str(uuid4())

        payload = {
            "tables": [
                {
                    "id": source_table_a_id,
                    "name": "Table A",
                    "fields": [
                        {
                            "id": source_a_title_id,
                            "name": "标题",
                            "field_type": "text",
                            "is_primary": True,
                            "order": 0,
                        },
                        {
                            "id": source_a_link_id,
                            "name": "关联B",
                            "field_type": "link",
                            "order": 1,
                            "config": {
                                "foreignTableId": source_table_b_id,
                                "lookupFieldId": source_b_title_id,
                                "filterByViewId": source_b_view_id,
                                "visibleFieldIds": [source_b_title_id],
                                "filter": {
                                    "conjunction": "and",
                                    "conditions": [
                                        {
                                            "fieldId": source_b_title_id,
                                            "operator": "is_not_empty",
                                        }
                                    ],
                                },
                            },
                        },
                        {
                            "id": source_a_formula_id,
                            "name": "标题公式",
                            "field_type": "formula",
                            "order": 2,
                            "config": {
                                "expression": f"CONCATENATE({{{source_a_title_id}}})",
                            },
                        },
                    ],
                    "records": [],
                    "views": [],
                },
                {
                    "id": source_table_b_id,
                    "name": "Table B",
                    "fields": [
                        {
                            "id": source_b_title_id,
                            "name": "名称",
                            "field_type": "text",
                            "is_primary": True,
                            "order": 0,
                        }
                    ],
                    "records": [],
                    "views": [
                        {
                            "id": source_b_view_id,
                            "name": "B筛选视图",
                            "view_type": "grid",
                            "order": 0,
                            "visible_fields": [source_b_title_id],
                            "field_order": [source_b_title_id],
                            "filters": [],
                            "sorts": [],
                            "groups": [],
                            "config": {},
                        }
                    ],
                },
            ],
            "metadata": {"format": "base_full"},
        }

        service = ImportService(user=self.user)
        result = service.import_space_from_json(
            space_id=self.space.id,
            json_content=json.dumps(payload, ensure_ascii=False),
            skip_errors=False,
        )

        self.assertEqual(result["created_tables"], 2)
        self.assertEqual(result["errors"], [])

        table_a = Table.objects.get(space_id=self.space.id, name="Table A")
        table_b = Table.objects.get(space_id=self.space.id, name="Table B")
        a_title = TableField.objects.get(table=table_a, name="标题")
        a_link = TableField.objects.get(table=table_a, name="关联B")
        a_formula = TableField.objects.get(table=table_a, name="标题公式")
        b_title = TableField.objects.get(table=table_b, name="名称")
        b_view = TableView.objects.get(table=table_b, name="B筛选视图")

        self.assertEqual(a_formula.field_type, "text")
        self.assertEqual(a_link.config.get("foreignTableId"), str(table_b.id))
        self.assertEqual(a_link.config.get("lookupFieldId"), str(b_title.id))
        self.assertEqual(a_link.config.get("filterByViewId"), str(b_view.id))
        self.assertEqual(a_link.config.get("visibleFieldIds"), [str(b_title.id)])
        self.assertEqual(
            a_link.config.get("filter", {}).get("conditions", [])[0].get("fieldId"),
            str(b_title.id),
        )
        self.assertEqual(
            a_formula.config.get("expression"),
            f"CONCATENATE({{{str(a_title.id)}}})",
        )
        self.assertNotIn(source_table_b_id, json.dumps(a_link.config, ensure_ascii=False))
        self.assertNotIn(source_b_title_id, json.dumps(a_link.config, ensure_ascii=False))
        self.assertNotIn(source_b_view_id, json.dumps(a_link.config, ensure_ascii=False))

    def test_json_import_project_base_full_supports_field_and_view_aliases(self):
        """测试项目级导入兼容外部表格导出 field/view 别名结构（type/options/lookupOptions/filter/sort/group）"""
        source_table_a_id = str(uuid4())
        source_table_b_id = str(uuid4())
        source_a_title_id = str(uuid4())
        source_a_link_id = str(uuid4())
        source_a_formula_id = str(uuid4())
        source_a_rollup_id = str(uuid4())
        source_a_view_id = str(uuid4())
        source_b_title_id = str(uuid4())
        source_b_num_id = str(uuid4())
        source_b_view_id = str(uuid4())

        payload = {
            "tables": [
                {
                    "id": source_table_a_id,
                    "name": "Alias Table A",
                    "fields": [
                        {
                            "id": source_a_title_id,
                            "name": "标题",
                            "type": "singleLineText",
                            "isPrimary": True,
                            "notNull": True,
                            "order": 0,
                        },
                        {
                            "id": source_a_link_id,
                            "name": "关联B",
                            "type": "link",
                            "order": 1,
                            "options": {
                                "foreignTableId": source_table_b_id,
                                "lookupFieldId": source_b_title_id,
                                "filterByViewId": source_b_view_id,
                                "visibleFieldIds": [source_b_title_id],
                            },
                        },
                        {
                            "id": source_a_formula_id,
                            "name": "标题公式",
                            "type": "formula",
                            "order": 2,
                            "options": {
                                "expression": f"CONCATENATE({{{source_a_title_id}}})",
                            },
                        },
                        {
                            "id": source_a_rollup_id,
                            "name": "B求和",
                            "type": "conditionalRollup",
                            "order": 3,
                            "options": {
                                "expression": "sum({values})",
                            },
                            "lookupOptions": {
                                "foreignTableId": source_table_b_id,
                                "lookupFieldId": source_b_num_id,
                                "linkFieldId": source_a_link_id,
                            },
                        },
                    ],
                    "records": [],
                    "views": [
                        {
                            "id": source_a_view_id,
                            "name": "A 视图",
                            "type": "grid",
                            "visible_fields": [source_a_title_id, source_a_link_id],
                            "field_order": [source_a_title_id, source_a_link_id, source_a_formula_id],
                            "filter": [
                                {"fieldId": source_a_title_id, "operator": "is_not_empty"},
                            ],
                            "sort": [
                                {"field": source_a_title_id, "direction": "asc"},
                            ],
                            "group": [
                                {"field": source_a_link_id},
                            ],
                            "options": {
                                "columnWidths": {
                                    source_a_title_id: 260,
                                }
                            },
                            "columnMeta": {
                                source_a_title_id: {"order": 0, "width": 260},
                                source_a_link_id: {"order": 1},
                                source_a_formula_id: {"order": 2},
                            },
                        }
                    ],
                    "defaultViewId": source_a_view_id,
                },
                {
                    "id": source_table_b_id,
                    "name": "Alias Table B",
                    "fields": [
                        {
                            "id": source_b_title_id,
                            "name": "名称",
                            "type": "singleLineText",
                            "isPrimary": True,
                            "order": 0,
                        },
                        {
                            "id": source_b_num_id,
                            "name": "数值",
                            "type": "number",
                            "order": 1,
                        },
                    ],
                    "records": [],
                    "views": [
                        {
                            "id": source_b_view_id,
                            "name": "B 视图",
                            "type": "grid",
                            "visible_fields": [source_b_title_id, source_b_num_id],
                            "field_order": [source_b_title_id, source_b_num_id],
                        }
                    ],
                    "defaultViewId": source_b_view_id,
                },
            ],
            "metadata": {"format": "base_full"},
        }

        service = ImportService(user=self.user)
        result = service.import_space_from_json(
            space_id=self.space.id,
            json_content=json.dumps(payload, ensure_ascii=False),
            skip_errors=False,
        )

        self.assertEqual(result["created_tables"], 2)
        self.assertEqual(result["errors"], [])

        table_a = Table.objects.get(space_id=self.space.id, name="Alias Table A")
        table_b = Table.objects.get(space_id=self.space.id, name="Alias Table B")

        a_title = TableField.objects.get(table=table_a, name="标题")
        a_link = TableField.objects.get(table=table_a, name="关联B")
        a_formula = TableField.objects.get(table=table_a, name="标题公式")
        a_rollup = TableField.objects.get(table=table_a, name="B求和")
        b_title = TableField.objects.get(table=table_b, name="名称")
        b_num = TableField.objects.get(table=table_b, name="数值")
        b_view = TableView.objects.get(table=table_b, name="B 视图")

        self.assertEqual(a_title.field_type, "text")
        self.assertEqual(a_link.field_type, "link")
        self.assertEqual(a_formula.field_type, "text")
        self.assertEqual(a_rollup.field_type, "text")

        self.assertEqual(a_link.config.get("foreignTableId"), str(table_b.id))
        self.assertEqual(a_link.config.get("lookupFieldId"), str(b_title.id))
        self.assertEqual(a_link.config.get("filterByViewId"), str(b_view.id))
        self.assertEqual(a_link.config.get("visibleFieldIds"), [str(b_title.id)])
        self.assertEqual(
            a_formula.config.get("expression"),
            f"CONCATENATE({{{str(a_title.id)}}})",
        )

        self.assertEqual(a_rollup.config.get("expression"), "sum({values})")
        self.assertEqual(
            a_rollup.config.get("lookupOptions", {}).get("foreignTableId"),
            str(table_b.id),
        )
        self.assertEqual(
            a_rollup.config.get("lookupOptions", {}).get("lookupFieldId"),
            str(b_num.id),
        )
        self.assertEqual(
            a_rollup.config.get("lookupOptions", {}).get("linkFieldId"),
            str(a_link.id),
        )

        a_view = TableView.objects.get(table=table_a, name="A 视图")
        self.assertEqual(a_view.filters[0]["fieldId"], str(a_title.id))
        self.assertEqual(a_view.sorts[0]["field"], str(a_title.id))
        self.assertEqual(a_view.groups[0]["field"], str(a_link.id))
        self.assertEqual(a_view.config["columnWidths"][str(a_title.id)], 260)
        self.assertEqual(str(table_a.default_view_id), str(a_view.id))

        table_a_result = next(
            item for item in result["table_results"] if item["table_id"] == str(table_a.id)
        )
        self.assertTrue(
            any("conditionalRollup" in str(message) for message in table_a_result["errors"])
        )

    def test_export_stats(self):
        """测试导出统计"""
        # 创建3条记录
        name_key = str(self.field_name.id)
        age_key = str(self.field_age.id)
        email_key = str(self.field_email.id)
        TableRecord.objects.bulk_create([
            TableRecord(
                table=self.table,
                data={name_key: f'用户{i}', age_key: 20 + i, email_key: f'user{i}@example.com'},
                created_by=self.user
            )
            for i in range(3)
        ])

        service = ExportService(user=self.user)
        stats = service.get_export_stats(table_id=self.table.id)

        self.assertIsNotNone(stats)
        self.assertEqual(stats['field_count'], 3)
        self.assertEqual(stats['record_count'], 3)
        self.assertIn('estimated_size', stats)
        self.assertIn('csv_kb', stats['estimated_size'])
        self.assertIn('excel_kb', stats['estimated_size'])
        self.assertIn('json_kb', stats['estimated_size'])
        self.assertIn('pdf_kb', stats['estimated_size'])

    def test_import_template_generation(self):
        """测试导入模板生成"""
        service = ImportService(user=self.user)
        template = service.get_import_template(table_id=self.table.id)

        self.assertIsNotNone(template)
        self.assertTrue(template.startswith('\ufeff姓名,年龄,邮箱'))
        self.assertTrue(template.encode('utf-8').startswith(b'\xef\xbb\xbf'))
        self.assertIn('姓名', template)
        self.assertIn('年龄', template)
        self.assertIn('邮箱', template)
        # 应该包含示例数据
        self.assertIn('示例文本1', template)
        self.assertIn('示例文本2', template)

    def test_import_template_json_generation(self):
        """JSON 导入模板：至少 2 行对象数组，key=显示名，便于识别表头与分行"""
        service = ImportService(user=self.user)
        template = service.get_import_template(table_id=self.table.id, format='json')

        payload = json.loads(template)
        self.assertIsInstance(payload, list)
        self.assertGreaterEqual(len(payload), 2)
        self.assertEqual(payload[0].get('姓名'), '示例文本1')
        self.assertEqual(payload[0].get('年龄'), 123)
        self.assertEqual(payload[1].get('姓名'), '示例文本2')
        self.assertEqual(payload[1].get('年龄'), 456)
        self.assertIn('邮箱', payload[0])

    def test_import_template_rejects_unknown_format(self):
        service = ImportService(user=self.user)
        with self.assertRaises(ValueError):
            service.get_import_template(table_id=self.table.id, format='xml')

    def test_validate_record_rejects_zero_field_match(self):
        """#5671：英文字段 key 全部未命中时校验失败（避免静默空写）"""
        from apps.tabdata.services.record_service import RecordService

        service = RecordService(user=self.user)
        ok, err = service._validate_record_data(
            self.table.id,
            {'title': '融资快讯', 'company_name': '示例科技'},
        )
        self.assertFalse(ok)
        self.assertIsNotNone(err)
        self.assertIn('无有效字段匹配', err)
        self.assertIn('title', err)

    def test_validate_record_partial_unknown_keys_collectable(self):
        """部分 key 命中时校验通过，未知 key 可被收集为 warnings"""
        from apps.tabdata.services.record_service import RecordService

        service = RecordService(user=self.user)
        data = {'姓名': '用户1', '年龄': 20, 'ghost_field': '应被忽略'}
        ok, err = service._validate_record_data(self.table.id, data)
        self.assertTrue(ok, err)
        fields = list(
            TableField.objects.filter(table=self.table, is_deleted=False)
        )
        unknown = service._collect_unknown_field_keys(data, fields)
        self.assertEqual(unknown, ['ghost_field'])

    def test_open_create_rejects_all_unknown_fields_without_write(self):
        """Open API：全部未知字段时 400，且不调用 create_record"""
        from django.test import RequestFactory
        from unittest.mock import patch
        from apps.tabdata.api_open_impl.record_impl import create_record_impl
        from apps.tabdata.api_open_schemas import OpenCreateRecordBody

        body = OpenCreateRecordBody(
            fields={'title': '幽灵', 'company_name': '空写'},
            field_key_type='name',
        )
        request = RequestFactory().post('/fake')
        request.auth = self.user
        request.api_token = None

        with patch(
            'apps.tabdata.api_open_impl.record_impl.RecordService.create_record'
        ) as mock_create:
            resp = create_record_impl(request, self.table.id, body)

        self.assertEqual(resp.status_code, 400)
        data = json.loads(resp.content)
        self.assertFalse(data.get('success', True))
        self.assertIn('无有效字段匹配', json.dumps(data, ensure_ascii=False))
        mock_create.assert_not_called()

    def _partial_bad_csv(self):
        """1 行坏数据（列数不匹配，触发整行 abort/skip）+ 5 行好数据。"""
        lines = ["姓名,年龄,邮箱", "坏数据,列数不匹配"]
        lines += [f"用户{i},{20 + i},user{i}@example.com" for i in range(5)]
        return "\n".join(lines)

    def test_csv_import_abort_on_error_does_not_report_false_success(self):
        """：skip_errors=False 遇坏行应在写库前中止，返回真实的 0/0，不虚报成功。"""
        service = ImportService(user=self.user)
        created, updated, errors = service.import_from_csv(
            table_id=self.table.id,
            file_content=self._partial_bad_csv(),
            skip_errors=False,
        )

        self.assertEqual(created, 0)
        self.assertEqual(updated, 0)
        self.assertTrue(errors)
        self.assertTrue(any('导入已中止' in e for e in errors))

        # DB 实际零写入，不应残留"部分导入"的记录
        self.assertEqual(
            TableRecord.objects.filter(table=self.table, is_deleted=False).count(), 0
        )

    def test_csv_import_skip_errors_writes_good_rows(self):
        """同一份坏数据在 skip_errors=True 下应正常写入好行、跳过坏行。"""
        service = ImportService(user=self.user)
        created, updated, errors = service.import_from_csv(
            table_id=self.table.id,
            file_content=self._partial_bad_csv(),
            skip_errors=True,
        )

        self.assertEqual(created, 5)
        self.assertEqual(updated, 0)
        self.assertTrue(errors)

        self.assertEqual(
            TableRecord.objects.filter(table=self.table, is_deleted=False).count(), 5
        )

    def test_csv_import_skips_blank_rows_without_error(self):
        """ 验收：数据区中间/末尾的全空行应被静默跳过，不中止导入、不报错。

        Excel/CSV 常在使用区内夹带空行；这些行不承载数据，
        既不该写入，也不该触发『没有可导入的有效字段』把整份导入拉挂。
        """
        csv_content = "\n".join([
            "姓名,年龄,邮箱",
            "张三,25,zhang@example.com",
            ",,",                       # 中间全空行
            "李四,30,li@example.com",
            ",,",                       # 末尾全空行
        ])

        service = ImportService(user=self.user)
        created, updated, errors = service.import_from_csv(
            table_id=self.table.id,
            file_content=csv_content,
            skip_errors=False,
        )

        self.assertEqual(created, 2)
        self.assertEqual(updated, 0)
        self.assertEqual(errors, [])
        self.assertEqual(
            TableRecord.objects.filter(table=self.table, is_deleted=False).count(), 2
        )


class ImportPreviewTestCase(TestCase):
    """导入预览功能测试"""

    databases = {'default', 'postgresql'}

    def setUp(self):
        """测试前准备"""
        self.user = User.objects.create_user(
            username='testuser2',
            email='test2@example.com',
            password='password123'
        )

        self.organization = Organization.objects.create(
            name='Test Organization 2',
            owner=self.user
        )

        OrganizationMember.objects.create(
            organization=self.organization,
            user=self.user,
            role='owner'
        )

        # ：Space 表已 DROP；Table.space_id 挂 Project.id
        self.space = Project.objects.create(
            name='Test Space 2',
            organization=self.organization,
        )

        self.table = Table.objects.create(
            space_id=self.space.id,
            organization_id=self.organization.id,
            name='Test Table 2',
            owner=self.user
        )

        # 创建字段
        TableField.objects.create(
            table=self.table,
            name='姓名',
            field_type='text',
            is_primary=True,
            order=0
        )

        TableField.objects.create(
            table=self.table,
            name='年龄',
            field_type='number',
            order=1
        )

    def test_import_preview_csv(self):
        """测试CSV导入预览"""
        csv_content = """姓名,年龄,邮箱
张三,25,zhang@example.com
李四,abc,li@example.com
王五,28,wang@example.com"""

        service = ImportService(user=self.user)
        result = service.preview_import(
            table_id=self.table.id,
            file_content=csv_content,
            file_type='csv',
            preview_rows=10
        )

        self.assertEqual(result['stats']['total_rows'], 3)
        self.assertEqual(result['stats']['preview_rows'], 3)
        self.assertEqual(result['stats']['field_count'], 3)
        self.assertEqual(len(result['preview_data']), 3)
        self.assertIsInstance(result['preview_data'][0], dict)
        self.assertIn('姓名', result['preview_data'][0])

        # 验证字段映射（数组格式，每个元素含 source/target/confidence/inferred_type）
        fm = result['field_mapping']
        self.assertEqual(len(fm), 3)
        sources = [m['source'] for m in fm]
        self.assertIn('姓名', sources)
        self.assertIn('年龄', sources)
        matched = [m for m in fm if m['confidence'] > 0]
        self.assertEqual(len(matched), 2)
        unmatched = [m for m in fm if m['confidence'] == 0]
        self.assertEqual(len(unmatched), 1)
        self.assertEqual(unmatched[0]['source'], '邮箱')

        # 验证错误检测（李四的年龄是abc，不是数字）
        self.assertGreater(len(result['validation_issues']), 0)
        self.assertIn('field', result['validation_issues'][0])
        self.assertIn('issue', result['validation_issues'][0])

    def test_import_preview_csv_2000_rows_supported(self):
        """2000 行 CSV 属于支持范围：预览 stats.total_rows 必须完整计数。"""
        import time

        rows = ['姓名,年龄,邮箱']
        rows.extend(f'用户{i},{20 + (i % 50)},user{i}@example.com' for i in range(2000))
        csv_content = '\n'.join(rows)

        service = ImportService(user=self.user)
        started = time.perf_counter()
        result = service.preview_import(
            table_id=self.table.id,
            file_content=csv_content,
            file_type='csv',
            preview_rows=10,
        )
        elapsed_ms = (time.perf_counter() - started) * 1000

        self.assertEqual(result['stats']['total_rows'], 2000)
        self.assertEqual(result['stats']['preview_rows'], 10)
        self.assertEqual(len(result['preview_data']), 10)
        # 基线记录：当前实现应远低于 5s；超时另开性能 issue，不与 TLS 故障混修。
        self.assertLess(elapsed_ms, 5000, f'2000-row preview too slow: {elapsed_ms:.1f}ms')

    def test_import_csv_2000_rows_not_silently_truncated(self):
        """同步导入不得在 1000 行静默截断：2000 行应全部创建。"""
        rows = ['姓名,年龄,邮箱']
        rows.extend(f'用户{i},{20 + (i % 50)},user{i}@example.com' for i in range(2000))
        csv_content = '\n'.join(rows)

        service = ImportService(user=self.user)
        created, updated, errors = service.import_from_csv(
            table_id=self.table.id,
            file_content=csv_content,
            skip_errors=True,
            auto_create_missing_fields=True,
        )

        self.assertEqual(created, 2000, msg=f'errors={errors}')
        self.assertEqual(updated, 0)
        self.assertFalse(
            any('截断' in str(err) for err in errors),
            msg=f'2000 行不应触发截断警告: {errors}',
        )


class ExportServiceExtendedTestCase(TestCase):
    """ExportService 导出核心路径测试"""

    databases = {'default', 'postgresql'}

    @staticmethod
    def _tiny_png_bytes():
        return ExportServiceExtendedTestCase._png_bytes(1, 1)

    @staticmethod
    def _png_bytes(width, height):
        from PIL import Image

        image = Image.new('RGB', (width, height), 'red')
        buffer = io.BytesIO()
        image.save(buffer, format='PNG')
        return buffer.getvalue()

    def setUp(self):
        self.user = User.objects.create_user(
            username='export_testuser',
            email='export_test@example.com',
            password='password123'
        )

        self.organization = Organization.objects.create(
            name='Export Test Organization',
            owner=self.user
        )

        OrganizationMember.objects.create(
            organization=self.organization,
            user=self.user,
            role='owner'
        )

        self.space = Space.objects.create(
            organization=self.organization,
            name='Export Test Space',
            description='For export tests',
        )

        self.table = Table.objects.create(
            space_id=self.space.id,
            organization_id=self.space.organization_id,
            name='Export Test Table',
            description='Export test table',
            owner=self.user
        )

        self.field_name = TableField.objects.create(
            table=self.table,
            name='姓名',
            field_type='text',
            is_primary=True,
            order=0
        )

        self.field_age = TableField.objects.create(
            table=self.table,
            name='年龄',
            field_type='number',
            order=1
        )

        self.field_email = TableField.objects.create(
            table=self.table,
            name='邮箱',
            field_type='email',
            order=2
        )

        name_key = str(self.field_name.id)
        age_key = str(self.field_age.id)
        email_key = str(self.field_email.id)
        TableRecord.objects.bulk_create([
            TableRecord(
                table=self.table,
                data={name_key: '张三', age_key: 25, email_key: 'zhang@example.com'},
                created_by=self.user
            ),
            TableRecord(
                table=self.table,
                data={name_key: '李四', age_key: 30, email_key: 'li@example.com'},
                created_by=self.user
            ),
        ])

    def test_excel_export_basic(self):
        """Excel 导出：返回有效 xlsx，含表头和 2 行数据"""
        import openpyxl

        service = ExportService(user=self.user)
        result = service.export_to_excel(table_id=self.table.id, include_headers=True)

        self.assertIsInstance(result, bytes)
        self.assertGreater(len(result), 0)

        wb = openpyxl.load_workbook(io.BytesIO(result))
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        self.assertGreaterEqual(len(rows), 3)
        headers = rows[0]
        self.assertIn('姓名', headers)
        self.assertIn('年龄', headers)
        self.assertIn('邮箱', headers)

    def test_excel_export_embeds_data_url_image_field(self):
        """Excel 导出：attachment 字段中的 data URL 图片会写入 workbook drawing。"""
        import openpyxl

        image_field = TableField.objects.create(
            table=self.table,
            name='图片',
            field_type='attachment',
            order=3,
        )
        image_bytes = self._png_bytes(640, 480)
        image_data_url = 'data:image/png;base64,' + base64.b64encode(image_bytes).decode('ascii')
        TableRecord.objects.create(
            table=self.table,
            data={
                str(self.field_name.id): '带图记录',
                str(image_field.id): [
                    {
                        'name': 'inline.png',
                        'url': image_data_url,
                        'size': len(image_bytes),
                        'mime_type': 'image/png',
                    },
                    {
                        'name': 'notes.pdf',
                        'url': 'https://assets.example.test/notes.pdf',
                        'size': 1024,
                        'mime_type': 'application/pdf',
                    },
                ],
            },
            created_by=self.user,
        )

        service = ExportService(user=self.user)
        result = service.export_to_excel(table_id=self.table.id, include_headers=True)

        wb = openpyxl.load_workbook(io.BytesIO(result))
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        self.assertGreaterEqual(len(ws._images), 1)
        self.assertEqual(rows[-1][3], 'inline.png, notes.pdf')
        self.assertGreater(ws.row_dimensions[4].height, 140)
        self.assertGreaterEqual(ws.column_dimensions['D'].width, 35)
        with zipfile.ZipFile(io.BytesIO(result)) as archive:
            media_names = [name for name in archive.namelist() if name.startswith('xl/media/')]
            self.assertEqual(len(media_names), 1)
            from PIL import Image

            with Image.open(io.BytesIO(archive.read(media_names[0]))) as embedded:
                self.assertEqual(embedded.size, (640, 480))
            drawing_names = [name for name in archive.namelist() if name.startswith('xl/drawings/drawing')]
            self.assertEqual(len(drawing_names), 1)
            drawing_xml = archive.read(drawing_names[0]).decode('utf-8')
            self.assertIn('<rowOff>190500</rowOff>', drawing_xml)
            self.assertIn('cx="2286000"', drawing_xml)
            self.assertIn('cy="1714500"', drawing_xml)

    def test_excel_export_embeds_all_images_in_attachment_cell(self):
        """Excel 导出：同一 attachment 单元格的多张图片不应漏导，也不应锚到其它行列。"""
        import openpyxl

        image_field = TableField.objects.create(
            table=self.table,
            name='图片',
            field_type='attachment',
            order=3,
        )
        image_values = []
        for index in range(4):
            image_bytes = self._png_bytes(640, 480)
            image_values.append({
                'name': f'inline-{index}.png',
                'url': 'data:image/png;base64,' + base64.b64encode(image_bytes).decode('ascii'),
                'size': len(image_bytes),
                'mime_type': 'image/png',
            })
        TableRecord.objects.create(
            table=self.table,
            data={
                str(self.field_name.id): '周日',
                str(image_field.id): image_values,
            },
            created_by=self.user,
        )

        service = ExportService(user=self.user)
        result = service.export_to_excel(table_id=self.table.id, include_headers=True)

        wb = openpyxl.load_workbook(io.BytesIO(result))
        ws = wb.active
        self.assertEqual(len(ws._images), 4)
        self.assertGreater(ws.row_dimensions[4].height, 140)
        with zipfile.ZipFile(io.BytesIO(result)) as archive:
            media_names = [name for name in archive.namelist() if name.startswith('xl/media/')]
            self.assertEqual(len(media_names), 4)
            drawing_names = [name for name in archive.namelist() if name.startswith('xl/drawings/drawing')]
            self.assertEqual(len(drawing_names), 1)
            drawing_xml = archive.read(drawing_names[0]).decode('utf-8')

        self.assertEqual(drawing_xml.count('<col>3</col>'), 4)
        self.assertEqual(drawing_xml.count('<row>3</row>'), 4)

    @patch('apps.tabdata.services.export_service.get_oss_service')
    def test_excel_export_embeds_oss_image_field(self, mock_get_oss_service):
        """Excel 导出：带 OSS key 的附件图片会下载并写入 workbook drawing。"""
        import openpyxl

        image_bytes = self._tiny_png_bytes()
        mock_oss_service = mock_get_oss_service.return_value
        mock_oss_service.download_file.return_value = {
            'success': True,
            'data': {
                'content': image_bytes,
                'content_type': 'image/png',
            },
        }
        image_field = TableField.objects.create(
            table=self.table,
            name='附件图片',
            field_type='attachment',
            order=3,
        )
        TableRecord.objects.create(
            table=self.table,
            data={
                str(self.field_name.id): 'OSS 图片记录',
                str(image_field.id): [
                    {
                        'name': 'oss.png',
                        'url': 'https://assets.example.test/oss.png',
                        'size': len(image_bytes),
                        'mime_type': 'image/png',
                        'key': 'tabdata/export/oss.png',
                    }
                ],
            },
            created_by=self.user,
        )

        service = ExportService(user=self.user)
        result = service.export_to_excel(table_id=self.table.id, include_headers=True)

        wb = openpyxl.load_workbook(io.BytesIO(result))
        ws = wb.active
        self.assertGreaterEqual(len(ws._images), 1)
        mock_oss_service.download_file.assert_called_once_with('tabdata/export/oss.png')

    @patch('apps.tabdata.services.export_service.get_oss_service')
    @override_settings(
        ALIYUN_OSS_BUCKET_NAME='example-assets',
        ALIYUN_OSS_ENDPOINT='oss-cn-wuhan-lr.aliyuncs.com',
        ASSET_PUBLIC_DOMAIN='',
        ALIYUN_OSS_CDN_DOMAIN='',
    )
    def test_excel_export_embeds_trusted_oss_url_image_field(self, mock_get_oss_service):
        """Excel 导出：历史附件只有自家 OSS URL 时，仍能反解 object key 后嵌入图片。"""
        import openpyxl

        image_bytes = self._tiny_png_bytes()
        mock_oss_service = mock_get_oss_service.return_value
        mock_oss_service.download_file.return_value = {
            'success': True,
            'data': {
                'content': image_bytes,
                'content_type': 'image/png',
            },
        }
        image_field = TableField.objects.create(
            table=self.table,
            name='历史附件图片',
            field_type='attachment',
            order=3,
        )
        TableRecord.objects.create(
            table=self.table,
            data={
                str(self.field_name.id): '仅 URL 图片记录',
                str(image_field.id): [
                    {
                        'name': 'legacy.png',
                        'url': 'https://example-assets.oss-cn-wuhan-lr.aliyuncs.com/tabdata/export/legacy.png',
                        'size': len(image_bytes),
                        'mime_type': 'image/png',
                    }
                ],
            },
            created_by=self.user,
        )

        service = ExportService(user=self.user)
        result = service.export_to_excel(table_id=self.table.id, include_headers=True)

        wb = openpyxl.load_workbook(io.BytesIO(result))
        ws = wb.active
        self.assertGreaterEqual(len(ws._images), 1)
        mock_oss_service.download_file.assert_called_once_with('tabdata/export/legacy.png')

    @patch('apps.tabdata.services.export_service.get_oss_service')
    def test_excel_export_ignores_untrusted_external_image_url(self, mock_get_oss_service):
        """Excel 导出：任意外部 URL 不被后端主动抓取，避免导出路径形成 SSRF。"""
        image = {
            'name': 'external.png',
            'url': 'https://example.invalid/external.png',
            'size': 1024,
            'mime_type': 'image/png',
        }

        from apps.tabdata.services.export_service import _build_excel_image_from_file_field

        self.assertIsNone(_build_excel_image_from_file_field([image], {}))
        mock_get_oss_service.assert_not_called()

    def test_excel_export_number_field(self):
        """Excel 导出：age 为空字符串时不崩溃"""
        name_key = str(self.field_name.id)
        age_key = str(self.field_age.id)
        email_key = str(self.field_email.id)
        TableRecord.objects.create(
            table=self.table,
            data={name_key: '空年龄', age_key: '', email_key: 'empty@example.com'},
            created_by=self.user
        )

        service = ExportService(user=self.user)
        result = service.export_to_excel(table_id=self.table.id, include_headers=True)

        self.assertIsInstance(result, bytes)
        self.assertGreater(len(result), 0)

    def test_excel_export_matches_grid_number_and_date_display(self):
        """Excel 导出：数字/日期格式应和表格默认显示一致。"""
        import openpyxl

        date_field = TableField.objects.create(
            table=self.table,
            name='日期',
            field_type='date',
            order=3,
        )
        record = TableRecord.objects.create(
            table=self.table,
            data={
                str(self.field_age.id): 28,
                str(date_field.id): '2026-06-21',
            },
            created_by=self.user,
        )

        service = ExportService(user=self.user)
        result = service.export_to_excel(
            table_id=self.table.id,
            record_ids=[record.id],
            field_ids=[self.field_age.id, date_field.id],
            include_headers=True,
        )

        wb = openpyxl.load_workbook(io.BytesIO(result))
        ws = wb.active
        self.assertEqual(ws['A2'].value, 28)
        self.assertEqual(ws['A2'].number_format, 'General')
        self.assertEqual(ws['B2'].value, '2026/6/21')

    def test_excel_export_respects_custom_date_formatting_with_time(self):
        """Excel 导出：date 字段应遵循日期和时间格式配置。"""
        import openpyxl

        date_field = TableField.objects.create(
            table=self.table,
            name='自定义日期',
            field_type='date',
            config={
                'formatting': {
                    'date': 'D/M/YYYY',
                    'time': 'None',
                    'timeZone': 'Asia/Shanghai',
                }
            },
            order=3,
        )
        time_field = TableField.objects.create(
            table=self.table,
            name='纽约时间',
            field_type='date',
            config={
                'formatting': {
                    'date': 'YYYY-MM-DD',
                    'time': 'hh:mm A',
                    'timeZone': 'America/New_York',
                }
            },
            order=4,
        )
        record = TableRecord.objects.create(
            table=self.table,
            data={
                str(date_field.id): '2026-03-07',
                str(time_field.id): '2026-03-07T19:22:00Z',
            },
            created_by=self.user,
        )

        service = ExportService(user=self.user)
        result = service.export_to_excel(
            table_id=self.table.id,
            record_ids=[record.id],
            field_ids=[date_field.id, time_field.id],
            include_headers=True,
        )

        wb = openpyxl.load_workbook(io.BytesIO(result))
        ws = wb.active
        self.assertEqual(ws['A2'].value, '7/3/2026')
        self.assertEqual(ws['B2'].value, '2026-03-07 02:22 PM')

    def test_excel_export_keeps_midnight_time_for_date_with_time(self):
        """Excel 导出：启用时间的 date 不应把 00:00 误判为无时间。"""
        import openpyxl

        time_field = TableField.objects.create(
            table=self.table,
            name='午夜时间',
            field_type='date',
            config={
                'formatting': {
                    'date': 'YYYY/M/D',
                    'time': 'HH:mm',
                    'timeZone': 'UTC',
                }
            },
            order=3,
        )
        record = TableRecord.objects.create(
            table=self.table,
            data={str(time_field.id): '2026-03-07T00:00:00Z'},
            created_by=self.user,
        )

        service = ExportService(user=self.user)
        result = service.export_to_excel(
            table_id=self.table.id,
            record_ids=[record.id],
            field_ids=[time_field.id],
            include_headers=True,
        )

        wb = openpyxl.load_workbook(io.BytesIO(result))
        ws = wb.active
        self.assertEqual(ws['A2'].value, '2026/3/7 00:00')

    def test_pdf_export_basic(self):
        """PDF 导出：返回非空 bytes，以 %PDF 开头"""
        service = ExportService(user=self.user)
        result = service.export_to_pdf(table_id=self.table.id)

        self.assertIsInstance(result, bytes)
        self.assertGreater(len(result), 0)
        self.assertTrue(result[:5].startswith(b'%PDF'))

    def test_pdf_export_embeds_image_field_without_image_names(self):
        """PDF 导出：图片字段嵌入图片，不把图片文件名当列表文本。"""
        image_field = TableField.objects.create(
            table=self.table,
            name='图片',
            field_type='attachment',
            order=3,
        )
        image_bytes = self._png_bytes(640, 480)
        TableRecord.objects.create(
            table=self.table,
            data={
                str(self.field_name.id): '带图记录',
                str(image_field.id): [
                    {
                        'name': 'inline.png',
                        'url': 'data:image/png;base64,' + base64.b64encode(image_bytes).decode('ascii'),
                        'size': len(image_bytes),
                        'mime_type': 'image/png',
                    },
                    {
                        'name': 'notes.pdf',
                        'url': 'https://assets.example.test/notes.pdf',
                        'size': 1024,
                        'mime_type': 'application/pdf',
                    },
                ],
            },
            created_by=self.user,
        )

        service = ExportService(user=self.user)
        result = service.export_to_pdf(table_id=self.table.id, title='带图导出')

        self.assertTrue(result[:5].startswith(b'%PDF'))
        self.assertIn(b'/Subtype /Image', result)
        self.assertNotIn(b'inline.png', result)

    @patch('apps.tabdata.services.export_service.get_oss_service')
    def test_pdf_export_ignores_untrusted_external_image_url(self, mock_get_oss_service):
        """PDF 导出：外部图片 URL 不被后端下载，避免导出路径 SSRF。"""
        image_field = TableField.objects.create(
            table=self.table,
            name='图片',
            field_type='attachment',
            order=3,
        )
        TableRecord.objects.create(
            table=self.table,
            data={
                str(self.field_name.id): '外链图记录',
                str(image_field.id): [
                    {
                        'name': 'external.png',
                        'url': 'https://example.invalid/external.png',
                        'size': 1024,
                        'mime_type': 'image/png',
                    }
                ],
            },
            created_by=self.user,
        )

        service = ExportService(user=self.user)
        result = service.export_to_pdf(table_id=self.table.id, title='外链图导出')

        self.assertTrue(result[:5].startswith(b'%PDF'))
        self.assertNotIn(b'/Subtype /Image', result)
        self.assertNotIn(b'external.png', result)
        mock_get_oss_service.assert_not_called()

    def test_pdf_export_cjk(self):
        """CJK 字体探测：_get_cjk_font_name 返回 str"""
        from apps.tabdata.services.export_service import _get_cjk_font_name
        font_name = _get_cjk_font_name()
        self.assertIsInstance(font_name, str)

    @patch('apps.tabdata.services.export_service.MAX_EXPORT_ROWS', 3)
    def test_csv_export_max_rows(self):
        """CSV 导出：MAX_EXPORT_ROWS=3 时最多输出 3 条数据行（+1 表头）"""
        name_key = str(self.field_name.id)
        age_key = str(self.field_age.id)
        email_key = str(self.field_email.id)
        TableRecord.objects.bulk_create([
            TableRecord(
                table=self.table,
                data={name_key: f'额外用户{i}', age_key: 40 + i, email_key: f'extra{i}@example.com'},
                created_by=self.user
            )
            for i in range(5)
        ])

        service = ExportService(user=self.user)
        csv_content = service.export_to_csv(table_id=self.table.id, include_headers=True)

        lines = [line for line in csv_content.strip().split('\n') if line.strip()]
        self.assertLessEqual(len(lines), 4)

    @patch('apps.tabdata.services.export_service.MAX_EXPORT_ROWS', 3)
    def test_csv_streaming_max_rows(self):
        """CSV 流式导出：MAX_EXPORT_ROWS=3 时最多输出 3 条数据行（+1 表头）"""
        name_key = str(self.field_name.id)
        age_key = str(self.field_age.id)
        email_key = str(self.field_email.id)
        TableRecord.objects.bulk_create([
            TableRecord(
                table=self.table,
                data={name_key: f'流式用户{i}', age_key: 50 + i, email_key: f'stream{i}@example.com'},
                created_by=self.user
            )
            for i in range(5)
        ])

        service = ExportService(user=self.user)
        chunks = list(service.export_to_csv_streaming(table_id=self.table.id, include_headers=True))
        csv_content = ''.join(chunks)

        lines = [line for line in csv_content.strip().split('\n') if line.strip()]
        self.assertLessEqual(len(lines), 4)


class PDFExportRegressionTestCase(TestCase):
    """不依赖已下线 Space 模型的 PDF 导出回归测试。"""

    databases = {'default', 'postgresql'}

    def setUp(self):
        self.read_data_bulk_patcher = patch(
            'apps.tabdata.services.export_service.read_data_bulk',
        )
        self.read_data_bulk_patcher.start()
        self.addCleanup(self.read_data_bulk_patcher.stop)
        self.user = User.objects.create_user(
            username='pdf_export_regression',
            email='pdf-export-regression@example.com',
            password='password123',
            nickname='当前成员姓名',
        )
        self.organization = Organization.objects.create(
            name='PDF Export Regression Organization',
            owner=self.user,
        )
        OrganizationMember.objects.create(
            organization=self.organization,
            user=self.user,
            role='owner',
        )
        self.table = Table.objects.create(
            organization_id=self.organization.id,
            space_id=None,
            name='PDF Export Regression Table',
            owner=self.user,
        )
        self.primary_field = TableField.objects.create(
            table=self.table,
            name='问题描述',
            field_type='text',
            is_primary=True,
            order=0,
        )

    def test_member_names_use_current_and_historical_identity_with_safe_fallback(self):
        """用户类字段输出当前/历史姓名，未知成员不泄露完整 UUID。"""
        from django.utils import timezone
        from apps.tabdata.services import export_service as export_module

        departed_user_id = uuid4()
        unknown_user_id = uuid4()
        OrganizationMemberIdentitySnapshot.objects.create(
            organization=self.organization,
            user_id=departed_user_id,
            display_name='离开时姓名',
            left_at=timezone.now(),
        )
        member_field = TableField.objects.create(
            table=self.table,
            name='指派人',
            field_type='user',
            order=1,
        )
        creator_field = TableField.objects.create(
            table=self.table,
            name='提交人',
            field_type='created_by',
            order=2,
        )
        record = TableRecord.objects.create(
            table=self.table,
            data={
                str(member_field.id): [
                    str(self.user.id),
                    {'id': str(departed_user_id), 'name': '过期嵌入姓名'},
                    str(unknown_user_id),
                ],
                str(creator_field.id): str(self.user.id),
            },
            created_by=self.user,
        )

        captured_tables = []
        real_pdf_table = export_module.PDFTable

        def capture_table(*args, **kwargs):
            captured_tables.append(args[0])
            return real_pdf_table(*args, **kwargs)

        with (
            patch.object(export_module, 'read_data_bulk'),
            patch.object(export_module, 'PDFTable', side_effect=capture_table),
        ):
            result = ExportService(user=self.user).export_to_pdf(
                table_id=self.table.id,
                record_ids=[record.id],
                field_ids=[member_field.id, creator_field.id],
            )

        self.assertTrue(result[:5].startswith(b'%PDF'))
        exported_text = ' | '.join(
            cell.getPlainText()
            for table_data in captured_tables
            for row in table_data[1:]
            for cell in row
        )
        self.assertIn('当前成员姓名', exported_text)
        # 离组成员保留离开时的姓名并标注状态，读者才不会以为他还在岗
        self.assertIn('离开时姓名（已离职）', exported_text)
        self.assertNotIn('过期嵌入姓名', exported_text)
        # 查不到的 ID 说「未知」，不再断言他"离开过"，也不再露出 ID 片段
        self.assertIn('未知', exported_text)
        self.assertNotIn('已离开成员', exported_text)
        self.assertNotIn(str(unknown_user_id)[:8], exported_text)
        self.assertNotIn(str(self.user.id), exported_text)
        self.assertNotIn(str(departed_user_id), exported_text)
        self.assertNotIn(str(unknown_user_id), exported_text)

    def test_wide_table_is_sliced_and_primary_column_repeats(self):
        """宽表按可读列宽横向分片，且每个分片保留主字段。"""
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.units import inch
        from apps.tabdata.services import export_service as export_module

        extra_fields = [
            TableField.objects.create(
                table=self.table,
                name=f'业务字段{index}',
                field_type='text',
                order=index + 1,
                width=150,
            )
            for index in range(18)
        ]
        fields = [self.primary_field, *extra_fields]
        record = TableRecord.objects.create(
            table=self.table,
            data={
                str(field.id): f'第{index}列需要保持可读的中文内容与连续文本'
                for index, field in enumerate(fields)
            },
            created_by=self.user,
        )

        captured_tables = []
        captured_title_styles = []
        real_pdf_table = export_module.PDFTable
        real_paragraph = export_module.Paragraph

        def capture_table(*args, **kwargs):
            captured_tables.append((args[0], kwargs['colWidths']))
            return real_pdf_table(*args, **kwargs)

        def capture_paragraph(text, style, *args, **kwargs):
            if '列分片' in text:
                captured_title_styles.append(style)
            return real_paragraph(text, style, *args, **kwargs)

        with (
            patch.object(export_module, 'read_data_bulk'),
            patch.object(export_module, 'PDFTable', side_effect=capture_table),
            patch.object(export_module, 'Paragraph', side_effect=capture_paragraph),
        ):
            result = ExportService(user=self.user).export_to_pdf(
                table_id=self.table.id,
                record_ids=[record.id],
                field_ids=[field.id for field in fields],
                orientation='landscape',
            )

        self.assertTrue(result[:5].startswith(b'%PDF'))
        self.assertGreater(len(captured_tables), 1)
        self.assertEqual(len(captured_title_styles), len(captured_tables))
        self.assertTrue(all(style.keepWithNext is False for style in captured_title_styles))
        available_width = landscape(A4)[0] - 2 * 0.45 * inch
        exported_non_primary_headers = []
        for table_data, column_widths in captured_tables:
            headers = [cell.getPlainText() for cell in table_data[0]]
            self.assertEqual(headers[0], self.primary_field.name)
            self.assertLessEqual(sum(column_widths), available_width + 0.01)
            self.assertGreaterEqual(min(column_widths), 0.8 * inch)
            exported_non_primary_headers.extend(headers[1:])

        self.assertEqual(exported_non_primary_headers, [field.name for field in extra_fields])

    def test_file_fields_are_embedded_as_pdf_attachments_with_safe_unique_names(self):
        import fitz

        attachment_field = TableField.objects.create(
            table=self.table,
            name='附件',
            field_type='attachment',
            order=1,
        )
        first_content = b'first attachment content'
        second_content = b'second attachment content'
        record = TableRecord.objects.create(
            table=self.table,
            data={
                str(self.primary_field.id): '带附件记录',
                str(attachment_field.id): [
                    {
                        'name': '../report.txt',
                        'url': 'data:text/plain;base64,' + base64.b64encode(first_content).decode('ascii'),
                        'size': len(first_content),
                        'mime_type': 'text/plain',
                    },
                    {
                        'name': 'report.txt',
                        'url': 'data:text/plain;base64,' + base64.b64encode(second_content).decode('ascii'),
                        'size': len(second_content),
                        'mime_type': 'text/plain',
                    },
                ],
            },
            created_by=self.user,
        )

        result = ExportService(user=self.user).export_to_pdf(
            table_id=self.table.id,
            record_ids=[record.id],
            field_ids=[self.primary_field.id, attachment_field.id],
        )

        document = fitz.open(stream=result, filetype='pdf')
        self.assertEqual(document.embfile_get('report.txt'), first_content)
        self.assertEqual(document.embfile_get('report (2).txt'), second_content)
        self.assertNotIn('../report.txt', document.embfile_names())

    def test_image_attachment_keeps_page_preview_and_embedded_original(self):
        import fitz

        image_field = TableField.objects.create(
            table=self.table,
            name='图片',
            field_type='attachment',
            order=1,
        )
        image_bytes = ExportServiceExtendedTestCase._png_bytes(640, 480)
        record = TableRecord.objects.create(
            table=self.table,
            data={
                str(image_field.id): [{
                    'name': 'preview.png',
                    'url': 'data:image/png;base64,' + base64.b64encode(image_bytes).decode('ascii'),
                    'size': len(image_bytes),
                    'mime_type': 'image/png',
                }],
            },
            created_by=self.user,
        )

        result = ExportService(user=self.user).export_to_pdf(
            table_id=self.table.id,
            record_ids=[record.id],
            field_ids=[image_field.id],
        )

        document = fitz.open(stream=result, filetype='pdf')
        self.assertIn(b'/Subtype /Image', result)
        self.assertEqual(document.embfile_get('preview.png'), image_bytes)

    @patch('apps.tabdata.services.export_service.get_oss_service')
    def test_oss_non_image_file_is_embedded_without_fetching_external_urls(self, mock_get_oss_service):
        import fitz

        attachment_field = TableField.objects.create(
            table=self.table,
            name='附件',
            field_type='attachment',
            order=1,
        )
        file_content = b'%PDF-1.4\nattachment payload'
        mock_get_oss_service.return_value.download_file.return_value = {
            'success': True,
            'data': {'content': file_content, 'content_type': 'application/pdf'},
        }
        record = TableRecord.objects.create(
            table=self.table,
            data={
                str(attachment_field.id): [
                    {
                        'name': 'spec.pdf',
                        'object_key': 'tabdata/attachments/spec.pdf',
                        'size': len(file_content),
                        'mime_type': 'application/pdf',
                    },
                    {
                        'name': 'external.txt',
                        'url': 'https://example.invalid/external.txt',
                        'size': 10,
                        'mime_type': 'text/plain',
                    },
                ],
            },
            created_by=self.user,
        )
        file_record = FileRecord.objects.create(
            file_name='spec.pdf',
            file_key='tabdata/attachments/spec.pdf',
            file_path='tabdata/attachments',
            file_size=len(file_content),
            file_type='document',
            mime_type='application/pdf',
            file_extension='pdf',
            file_hash='pdf-export-spec',
            bucket_name='test-bucket',
            organization_id=str(self.organization.id),
            status='completed',
        )
        AttachmentReference.objects.create(
            organization_id=self.organization.id,
            table=self.table,
            field=attachment_field,
            record=record,
            file_id=file_record.id,
            created_by=self.user,
        )

        result = ExportService(user=self.user).export_to_pdf(
            table_id=self.table.id,
            record_ids=[record.id],
            field_ids=[attachment_field.id],
        )

        document = fitz.open(stream=result, filetype='pdf')
        self.assertEqual(document.embfile_get('spec.pdf'), file_content)
        self.assertNotIn('external.txt', document.embfile_names())
        mock_get_oss_service.return_value.download_file.assert_called_once_with(
            'tabdata/attachments/spec.pdf',
        )

    @patch('apps.tabdata.services.export_service.get_oss_service')
    def test_pdf_export_does_not_download_unreferenced_or_cross_organization_object_key(
        self,
        mock_get_oss_service,
    ):
        import fitz

        attachment_field = TableField.objects.create(
            table=self.table,
            name='闄勪欢',
            field_type='attachment',
            order=1,
        )
        foreign_organization = Organization.objects.create(
            name='Foreign PDF Attachment Organization',
            owner=self.user,
        )
        foreign_file = FileRecord.objects.create(
            file_name='foreign-secret.txt',
            file_key='tabdata/foreign/secret.txt',
            file_path='tabdata/foreign',
            file_size=12,
            file_type='document',
            mime_type='text/plain',
            file_extension='txt',
            file_hash='pdf-export-foreign-secret',
            bucket_name='test-bucket',
            organization_id=str(foreign_organization.id),
            status='completed',
        )
        record = TableRecord.objects.create(
            table=self.table,
            data={str(attachment_field.id): [{
                'name': 'foreign-secret.txt',
                'file_id': str(foreign_file.id),
                'object_key': foreign_file.file_key,
            }]},
            created_by=self.user,
        )

        result = ExportService(user=self.user).export_to_pdf(
            table_id=self.table.id,
            record_ids=[record.id],
            field_ids=[attachment_field.id],
        )

        document = fitz.open(stream=result, filetype='pdf')
        self.assertNotIn('foreign-secret.txt', document.embfile_names())
        mock_get_oss_service.assert_not_called()

    @patch('apps.tabdata.services.export_service.get_oss_service')
    def test_pdf_export_rejects_trusted_oversized_file_before_oss_download(
        self,
        mock_get_oss_service,
    ):
        from apps.tabdata.services import export_service as export_module

        attachment_field = TableField.objects.create(
            table=self.table,
            name='Oversized attachment',
            field_type='attachment',
            order=1,
        )
        file_record = FileRecord.objects.create(
            file_name='oversized.bin',
            file_key='tabdata/attachments/oversized.bin',
            file_path='tabdata/attachments',
            file_size=export_module._PDF_ATTACHMENT_MAX_BYTES + 1,
            file_type='other',
            mime_type='application/octet-stream',
            file_extension='bin',
            file_hash='pdf-export-oversized',
            bucket_name='test-bucket',
            organization_id=str(self.organization.id),
            status='completed',
        )
        record = TableRecord.objects.create(
            table=self.table,
            data={str(attachment_field.id): [{
                'name': file_record.file_name,
                'file_id': str(file_record.id),
                'object_key': file_record.file_key,
            }]},
            created_by=self.user,
        )
        AttachmentReference.objects.create(
            organization_id=self.organization.id,
            table=self.table,
            field=attachment_field,
            record=record,
            file_id=file_record.id,
            created_by=self.user,
        )

        ExportService(user=self.user).export_to_pdf(
            table_id=self.table.id,
            record_ids=[record.id],
            field_ids=[attachment_field.id],
        )

        mock_get_oss_service.assert_not_called()

    @patch('apps.tabdata.services.export_service.get_oss_service')
    def test_pdf_export_embeds_same_referenced_file_only_once(self, mock_get_oss_service):
        import fitz

        attachment_field = TableField.objects.create(
            table=self.table,
            name='鍏辩敤闄勪欢',
            field_type='attachment',
            order=1,
        )
        file_content = b'shared attachment'
        file_record = FileRecord.objects.create(
            file_name='shared.txt',
            file_key='tabdata/attachments/shared.txt',
            file_path='tabdata/attachments',
            file_size=len(file_content),
            file_type='document',
            mime_type='text/plain',
            file_extension='txt',
            file_hash='pdf-export-shared',
            bucket_name='test-bucket',
            organization_id=str(self.organization.id),
            status='completed',
        )
        records = [
            TableRecord.objects.create(
                table=self.table,
                data={str(attachment_field.id): [{
                    'name': file_record.file_name,
                    'file_id': str(file_record.id),
                    'object_key': file_record.file_key,
                }]},
                created_by=self.user,
            )
            for _ in range(2)
        ]
        for record in records:
            AttachmentReference.objects.create(
                organization_id=self.organization.id,
                table=self.table,
                field=attachment_field,
                record=record,
                file_id=file_record.id,
                created_by=self.user,
            )
        mock_get_oss_service.return_value.download_file.return_value = {
            'success': True,
            'data': {'content': file_content, 'content_type': 'text/plain'},
        }

        result = ExportService(user=self.user).export_to_pdf(
            table_id=self.table.id,
            record_ids=[record.id for record in records],
            field_ids=[attachment_field.id],
        )

        document = fitz.open(stream=result, filetype='pdf')
        self.assertEqual(document.embfile_names(), ['shared.txt'])
        self.assertEqual(document.embfile_get('shared.txt'), file_content)
        mock_get_oss_service.return_value.download_file.assert_called_once_with(file_record.file_key)


if __name__ == '__main__':
    import django
    import os
    os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'config.settings')
    django.setup()

    from django.test import TestCase
    from django.test.runner import DiscoverRunner

    runner = DiscoverRunner(verbosity=2)
    runner.run_tests(['apps.tabdata.tests.test_import_export_enhanced'])
